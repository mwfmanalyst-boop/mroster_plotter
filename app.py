import io
import re
import json
import tempfile
from datetime import date, timedelta, datetime
from typing import Optional, Tuple, List, Dict

import numpy as np
import pandas as pd
import streamlit as st
import altair as alt
import base64

# Google Drive API (service account)
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

# DuckDB
import duckdb

# ===== Auth deps (SSO) =====
from streamlit_oauth import OAuth2Component

# ===== Password (no-bcrypt) helpers =====
import hashlib, hmac
def _sha256_hex(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()

def verify_local_password(entered_password: str, salt_hex: str, stored_sha256_hex: str) -> bool:
    """
    Compares SHA-256(salt + entered_password) with stored_sha256_hex in constant time.
    """
    if not salt_hex or not stored_sha256_hex:
        return False
    calc = _sha256_hex(salt_hex + entered_password).lower()
    return hmac.compare_digest(calc, stored_sha256_hex.lower())

# =========================
# Page + Global Config + CSS
# =========================
st.set_page_config(page_title="Center Shiftwise Web Dashboard", layout="wide")

# --- UI state ---
if "active_tab" not in st.session_state:
    st.session_state.active_tab = "Dashboard"  # "Dashboard" | "Plotter" | "Roster" | "Admin"
if "busy_roster" not in st.session_state:
    st.session_state.busy_roster = False

# Animated gradient background + card style
# ===== Theme / Global Styles =====
st.markdown("""
/* App shell spacing */
header, .block-container { scroll-behavior: smooth; }
.block-container { padding-top: 8px !important; }

/* Make the sidebar info group sticky-ish */
section[data-testid="stSidebar"] .st-emotion-cache-1gulkj5 {
    position: sticky; top: 12px;
}

<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class^="css"] {
    font-family: 'Inter', system-ui, Segoe UI, Roboto, Ubuntu, 'Helvetica Neue', Arial, 'Noto Sans' !important;
}

/* Animated main gradient */
.stApp {
    background: linear-gradient(115deg, #667eea 0%, #764ba2 52%, #ff7eb3 100%);
    background-size: 320% 320%;
    animation: gradientMove 13s ease infinite;
    min-height: 100vh;
    padding: 28px 16px 48px 16px;
}
@keyframes gradientMove {
    0%,100% {background-position: 0% 50%;}
    50% {background-position: 100% 50%;}
}

/* Modern tabs */
.stTabs [data-baseweb="tab-list"] {
    background: linear-gradient(90deg, #ff7eb3 0%, #48cae4 100%);
    border-radius: 20px !important;
    box-shadow: 0 4px 14px rgba(80,80,200,0.14);
    padding: 8px 6px;
    margin-bottom: 26px;
}
.stTabs [data-baseweb="tab"] {
    color: #fff !important;
    font-weight: 600;
    border-radius: 14px !important;
    margin: 0 2px !important;
    background: rgba(255,255,255,0.11);
    transition: background .16s, color .16s;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(90deg,#764ba2 0,#48cae4 100%) !important;
    color: #fff !important;
    box-shadow: 0 2px 12px rgba(80,80,200,0.12);
}

/* Floating cards and containers */
.card, .block-container, .stTabs [data-baseweb="tab-list"], [data-testid="stMetric"], [data-testid="stDataFrame"] div[role="grid"], section[data-testid="stSidebar"], .stForm {
    background: rgba(255,255,255,0.10) !important;
    border: 1px solid rgba(255,255,255,0.21) !important;
    border-radius: 22px !important;
    box-shadow: 0 10px 36px 0 rgba(31, 38, 135, 0.14) !important;
    backdrop-filter: blur(17px) !important;
    margin-bottom: 24px;
    position: relative;
    z-index: 2;
}

/* Emphasize headers */
h1, h2, h3, h4, h5, h6, .stMarkdown p, .stMarkdown li {
    color: #FFFFFF !important;
    text-shadow: 1px 2px 16px rgba(0,0,0,0.22), 0 1px 4px #764ba24d;
}

/* Buttons, modern gradients */
.stButton>button {
    border-radius: 20px !important;
    border: none !important;
    background: linear-gradient(95deg, #ff7eb3 0%, #48cae4 100%);
    color: #fff !important;
    font-weight: 600 !important;
    padding: 9px 36px !important;
    box-shadow: 0 4px 22px rgba(188,143,255,0.19);
    transition: background 0.3s, transform 0.18s;
}
.stButton>button:hover {
    background: linear-gradient(95deg, #48cae4 0%, #764ba2 100%);
    transform: translateY(-2px) scale(1.048);
}

/* Metrics & DataFrame tweaks */
[data-testid="stMetricValue"], [data-testid="stMetricDelta"] {
    color: #e0e8ff !important;
}
[data-testid="stDataFrame"] table thead th {
    backdrop-filter: blur(10px);
    background: rgba(118,75,162,0.18);
    color: #fff;
}
[data-testid="stDataFrame"] tbody tr:nth-child(odd) td {
    background: rgba(118,75,162,0.10);
}

/* Sidebar gradient glassmorphism */
section[data-testid="stSidebar"] {
    background: linear-gradient(135deg, #764ba2 0%, #48cae4 100%);
    opacity: 0.95;
    border-right: 1px solid rgba(255,255,255,0.16);
}

/* Input fields */
input, textarea {
    background: rgba(255,255,255,0.19) !important;
    border-radius: 12px !important;
    color: #232b5c !important;
    border: none !important;
    font-weight: 500 !important;
    box-shadow: 0 1px 10px #764ba226;
}

/* Card title bold */
.card h4, .card h3, .card h2, .card h1 {
    font-weight: 700;
    letter-spacing: 1px;
    color: #fff;
    background: linear-gradient(90deg,#ff7eb3 0,#48cae4 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
}

.block-container {
    max-width: 1400px;
    margin-top: 18px !important;
}
/* ==== KPI gradient cards ==== */
.kpi-card{
  display:flex; flex-direction:column; justify-content:flex-start; 
  border-radius:22px; padding:18px 22px; min-height:110px;
  border:1px solid rgba(255,255,255,0.22);
  box-shadow: 0 10px 36px rgba(31,38,135,0.18);
  color:#fff;
  backdrop-filter: blur(10px);
}
.kpi-label{
  font-size:14px; font-weight:700; letter-spacing:.3px; opacity:.92;
  margin-bottom:8px; text-shadow: 0 1px 6px rgba(0,0,0,.18);
}
.kpi-value{
  font-size:42px; line-height:1; font-weight:800;
  text-shadow: 0 2px 14px rgba(0,0,0,.28);
}

/* variants */
.kpi-purple{
  background: linear-gradient(135deg,#a78bfa 0%, #7c3aed 100%);
}
.kpi-yellow{
  background: linear-gradient(135deg,#fde68a 0%, #f59e0b 100%);
  color:#1f2937; /* darker text for contrast on yellow */
}
.kpi-green{
  background: linear-gradient(135deg,#86efac 0%, #22c55e 100%);
}
.kpi-red{
  background: linear-gradient(135deg,#fda4af 0%, #ef4444 100%);
}
.kpi-neutral{
  background: linear-gradient(135deg,#94a3b8 0%, #64748b 100%);
}

</style>
""", unsafe_allow_html=True)
# =========================
# Full-screen Overlay Spinner helper
# =========================
from contextlib import contextmanager

def _overlay_html(message: str) -> str:
    return f"""
    <style>
    .__busy_overlay__ {{
        position: fixed;
        z-index: 99999;
        inset: 0;
        background: rgba(0, 0, 0, 0.45);
        backdrop-filter: blur(2px);
        display: flex;
        align-items: center;
        justify-content: center;
    }}
    .__busy_card__ {{
        background: rgba(255,255,255,0.10);
        border: 1px solid rgba(255,255,255,0.25);
        border-radius: 20px;
        padding: 22px 26px;
        color: #fff;
        font-weight: 600;
        box-shadow: 0 10px 36px rgba(31,38,135,0.25);
    }}
    .__busy_spinner__ {{
        width: 28px; height: 28px; border-radius: 50%;
        border: 3px solid rgba(255,255,255,0.35);
        border-top-color: #fff;
        animation: spin 0.8s linear infinite;
        margin-right: 12px;
    }}
    @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
    </style>
    <div class="__busy_overlay__">
      <div class="__busy_card__" style="display:flex;align-items:center;">
        <div class="__busy_spinner__"></div>
        <div>{message}</div>
      </div>
    </div>
    """

@contextmanager
def overlay(message: str = "Processing…"):
    """
    Full-screen overlay that appears immediately during work in the with-block.
    Usage:
        with overlay("Loading data…"):
            do_heavy_work()
    """
    placeholder = st.empty()
    placeholder.markdown(_overlay_html(message), unsafe_allow_html=True)
    try:
        yield
    finally:
        placeholder.empty()

# =========================
# 🔐 Auth + RBAC + ACL
# =========================
def _get_auth_config():
    cfg = st.secrets.get("auth", {})
    sso_roles = cfg.get("sso_roles", {})
    google = st.secrets.get("oauth", {}).get("google", {})
    return {
        "password_enabled": bool(cfg.get("enable_password_login", True)),
        "sso_enabled": bool(cfg.get("enable_sso_login", True)),
        "password_users": cfg.get("password_users", []),
        "admin_domains": set(sso_roles.get("admin_domains", [])),
        "viewer_domains": set(sso_roles.get("viewer_domains", [])),
        "google": google,
    }

def _domain_of(email: str) -> str:
    return (email.split("@", 1)[-1] or "").lower().strip()

def _role_from_sso_email(email: str, admin_domains: set[str], viewer_domains: set[str]) -> str:
    dom = _domain_of(email)
    if any(d == "*" or d == dom for d in admin_domains):
        return "admin"
    if any(d == "*" or d == dom for d in viewer_domains):
        return "viewer"
    return "viewer"

def _password_login_form(pusers: list[dict]) -> dict | None:
    st.markdown("### 🔑 Login with Username & Password")
    with st.form("pw_login"):
        u = st.text_input("Username", key="auth_username")
        p = st.text_input("Password", type="password", key="auth_password")
        submitted = st.form_submit_button("Login")
    if not submitted:
        return None
    if not u or not p:
        st.warning("Enter both username and password.")
        return None
    for rec in pusers:
        if rec.get("username", "").strip().lower() == u.strip().lower():
            ok = verify_local_password(
                entered_password=p,
                salt_hex=rec.get("password_salt", ""),
                stored_sha256_hex=rec.get("password_sha256", ""),
            )
            if ok:
                return {
                    "name": rec.get("name") or u,
                    "email": rec.get("email") or f"{u}@local",
                    "role": rec.get("role","viewer"),
                    "auth": "password",
                }
            break
    st.error("Invalid username or password.")
    return None

def _sso_login_google(google_cfg: dict, admin_domains: set[str], viewer_domains: set[str]) -> dict | None:
    """
    Google SSO (robust across streamlit_oauth variants) that avoids STATE mismatch by:
      - using a stable, normalized redirect_uri that matches Google Console exactly
      - clearing ?code&state after handling the callback
    Returns: {"name", "email", "role", "auth":"sso"} or None
    """
    import base64, json, requests, time
    from urllib.parse import urlencode

    st.markdown("### 🔓 Login with Google (SSO)")

    # ---- Required config
    client_id     = (google_cfg.get("client_id") or "").strip()
    client_secret = (google_cfg.get("client_secret") or "").strip()
    authorize_url = (google_cfg.get("authorize_url") or "").strip()
    token_url     = (google_cfg.get("token_url") or "").strip()
    user_info_url = (google_cfg.get("user_info_url") or "").strip()
    redirect_uri  = (google_cfg.get("redirect_uri") or google_cfg.get("redirect_url") or "").strip()


    if not all([client_id, client_secret, authorize_url, token_url, user_info_url, redirect_uri]):
        st.error(
            "SSO is not fully configured. Need: client_id, client_secret, authorize_url, "
            "token_url, user_info_url, redirect_uri in st.secrets['oauth']['google']"
        )
        return None

    # Remember the redirect we are using this run (stays stable across reruns)
    st.session_state.setdefault("_oauth_redirect_uri", redirect_uri)

    # ---- Build the component (positional args to support older wheels)
    try:
        oauth2 = OAuth2Component(
            client_id, client_secret, authorize_url,
            token_url, token_url,                 # token + refresh token
            st.session_state["_oauth_redirect_uri"]
        )
    except TypeError as e:
        st.error(f"OAuth2Component init failed: {e}")
        return None

    # Helper: decode id_token payload without verification (good enough to read email/name)
    def _decode_jwt_payload(id_token: str) -> dict:
        try:
            parts = id_token.split(".")
            if len(parts) != 3:
                return {}
            pad = lambda s: s + "=" * (-len(s) % 4)
            return json.loads(base64.urlsafe_b64decode(pad(parts[1])).decode("utf-8"))
        except Exception:
            return {}

    def _role(email: str) -> str:
        return _role_from_sso_email(email, admin_domains, viewer_domains)

    # ---- Render the button / handle callback
    # Some builds require positional redirect here too.
    try:
        result = oauth2.authorize_button(
            "Sign in with Google",
            st.session_state["_oauth_redirect_uri"],
            "openid email profile",
            key="google_oauth_btn",
            use_container_width=True,
            extras_params={"prompt": "consent", "access_type": "offline", "response_type": "code"},
        )
    except Exception as e:
        st.error(f"authorize_button failed: {e}")
        return None

    # No click yet
    if not result:
        return None

    # ---- We have a callback (Google redirected back)
    # Different wheels return different shapes: {"token": {...}} or flat dict.
    token = result.get("token", result if isinstance(result, dict) else {}) or {}

    # Try userinfo with access_token first
    email = name = None
    access_token = token.get("access_token")
    id_token     = token.get("id_token")

    if access_token:
        try:
            r = requests.get(user_info_url, headers={"Authorization": f"Bearer {access_token}"}, timeout=10)
            if r.ok:
                info = r.json() if r.content else {}
                email = (info.get("email") or "").strip().lower()
                name = (info.get("name")
                        or f"{info.get('given_name','')} {info.get('family_name','')}".strip()
                        or email or "User")
        except Exception as e:
            st.info(f"Note: userinfo fetch failed, using id_token if present ({e})")

    # Fallback: decode id_token
    if not email and id_token:
        payload = _decode_jwt_payload(id_token)
        email = (payload.get("email") or "").strip().lower()
        name = (payload.get("name")
                or f"{payload.get('given_name','')} {payload.get('family_name','')}".strip()
                or email or "User")

    if not email:
        st.error("SSO error: Could not obtain user email from Google (no userinfo and no usable id_token).")
        return None

    # ---- Clear the ?code&state from the URL so a rerun doesn't re-process and fail STATE check
    try:
        # Streamlit ≥ 1.30
        st.query_params.clear()
    except Exception:
        # Back-compat
        try:
            st.experimental_set_query_params()
        except Exception:
            pass

    return {"name": name, "email": email, "role": _role(email), "auth": "sso"}

def _render_auth_gate():
    if "user" in st.session_state and st.session_state["user"]:
        return st.session_state["user"]

    cfg = _get_auth_config()
    tabs = []
    if cfg["password_enabled"]: tabs.append("User/Password")
    if cfg["sso_enabled"]: tabs.append("SSO (Google)")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.subheader("Welcome — please sign in")
    t = st.tabs(tabs or ["Sign-in disabled"])

    user = None; ti = 0
    if cfg["password_enabled"]:
        with t[ti]: user = _password_login_form(cfg["password_users"]) or user
        ti += 1
    if cfg["sso_enabled"]:
        with t[ti if len(t)>ti else -1]:
            user = _sso_login_google(cfg["google"], cfg["admin_domains"], cfg["viewer_domains"]) or user
    st.markdown('</div>', unsafe_allow_html=True)

    if user:
        st.session_state["user"] = user
        st.success(f"Hello, {user['name']}!")
        st.rerun()
    st.stop()

ACL_FILE_NAME = st.secrets.get("ACL_FILE_NAME", "center_acl.json")

@st.cache_resource
def _get_drive_service():
    SA_INFO = st.secrets.get("gcp_service_account", {})
    if not SA_INFO:
        st.error("Missing st.secrets['gcp_service_account']. Add your service account JSON.")
        raise RuntimeError("No service account in secrets")
    scopes = ['https://www.googleapis.com/auth/drive']
    creds = service_account.Credentials.from_service_account_info(SA_INFO, scopes=scopes)
    return build('drive', 'v3', credentials=creds, cache_discovery=False)

def _drive_resolve_shortcut(service, file_obj):
    if file_obj.get("mimeType") == "application/vnd.google-apps.shortcut":
        target_id = file_obj.get("shortcutDetails", {}).get("targetId")
        if target_id:
            return service.files().get(fileId=target_id,
                                       fields="id,name,parents,mimeType,driveId",
                                       supportsAllDrives=True).execute()
    return file_obj

def _drive_get_file_by_id(service, file_id: str):
    return service.files().get(
        fileId=file_id,
        fields="id,name,parents,mimeType,driveId,shortcutDetails",
        supportsAllDrives=True
    ).execute()

def _drive_find_file_id(service, name: str, folder_id: Optional[str]) -> Optional[str]:
    q = f"name = '{name}' and trashed = false"
    if folder_id:
        q = f"name = '{name}' and '{folder_id}' in parents and trashed = false"
    resp = service.files().list(
        q=q,
        fields="files(id,name,parents,mimeType,shortcutDetails,driveId)",
        pageSize=50,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
        corpora="allDrives",
    ).execute()
    files = resp.get("files", [])
    if not files:
        return None
    files = [_drive_resolve_shortcut(service, f) for f in files]
    if folder_id:
        for f in files:
            if folder_id in (f.get("parents") or []):
                return f["id"]
    return files[0]["id"]

def _drive_download_bytes(service, file_id: str) -> bytes:
    req = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()

def _drive_upload_bytes(service, name: str, data: bytes, mime: str, folder_id: Optional[str], file_id: Optional[str] = None) -> str:
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime, resumable=False)
    if file_id:
        file = service.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()
        return file["id"]
    meta = {"name": name}
    if folder_id:
        meta["parents"] = [folder_id]
    file = service.files().create(body=meta, media_body=media, fields="id", supportsAllDrives=True).execute()
    return file["id"]

@st.cache_data(ttl=120, show_spinner=False)
@st.cache_data(ttl=120, show_spinner=False)
def _drive_list_folder(folder_id: str) -> list[dict]:
    """
    Cached list of files in a Drive folder.
    Only accepts hashable parameters so Streamlit cache works.
    """
    service = _get_drive_service()  # this is @st.cache_resource in your code
    items = []
    page_token = None
    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            fields="nextPageToken, files(id,name,mimeType,shortcutDetails)",
            pageSize=200,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="allDrives"
        ).execute()
        for f in resp.get("files", []):
            items.append(_drive_resolve_shortcut(service, f))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items

@st.cache_data(ttl=300)
def _load_acl_from_drive_cached(folder_id: Optional[str]) -> dict:
    try:
        service = _get_drive_service()
        file_id = _drive_find_file_id(service, ACL_FILE_NAME, folder_id or None)
        if not file_id:
            return {}
        raw = _drive_download_bytes(service, file_id)
        return json.loads(raw.decode("utf-8"))
    except Exception:
        return {}

def _save_acl_to_drive(acl: dict, folder_id: Optional[str]):
    try:
        service = _get_drive_service()
        data = json.dumps(acl, indent=2).encode("utf-8")
        file_id = _drive_find_file_id(service, ACL_FILE_NAME, folder_id or None)
        _drive_upload_bytes(service, ACL_FILE_NAME, data, "application/json", folder_id or None, file_id)
        _load_acl_from_drive_cached.clear()
    except Exception as e:
        st.error(f"Failed saving ACL: {e}")

def _norm_token(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', (s or '').lower())

def resolve_allowed_centers_from_email(email: str, all_centers: List[str], acl: dict) -> tuple[list[str], bool]:
    if not email:
        return [], False
    e = email.lower()
    if "meesho" in e:
        return list(all_centers), True
    email_tok = _norm_token(email)
    allowed = set()
    for c in all_centers:
        ctok = _norm_token(c)
        if ctok and ctok in email_tok:
            allowed.add(c)
        editors = set(map(str.lower, (acl.get(c, {}).get("editors") or [])))
        if email.lower() in editors:
            allowed.add(c)
    return sorted(allowed), False

def user_can_edit_center(email: str, role: str, center: str, allowed_centers: list[str], has_full_access: bool) -> bool:
    if role != "admin":
        return False
    if has_full_access:
        return True
    return center in set(allowed_centers)

# =========================
# Secrets (read once)
# =========================
SA_INFO = st.secrets.get("gcp_service_account", {})
DRIVE_FOLDER_ID = st.secrets.get("DRIVE_FOLDER_ID", "").strip()
DUCKDB_FILE_NAME = st.secrets.get("DUCKDB_FILE_NAME", "").strip() or "cmb_delta.duckdb"
DUCKDB_FILE_ID = st.secrets.get("DUCKDB_FILE_ID", "").strip()  # optional but best
REQUESTED_FOLDER_ID = st.secrets.get("REQUESTED_FOLDER_ID", "").strip()
ROSTER_FOLDER_ID = st.secrets.get("ROSTER_FOLDER_ID", "").strip()

# =========================
# Authenticate user
# =========================
user = _render_auth_gate()

# =========================
# Parquet helpers (fallback store)
# =========================
@st.cache_data(ttl=300)
def load_parquet_from_drive(name: str) -> pd.DataFrame:
    service = _get_drive_service()
    folder_id = DRIVE_FOLDER_ID or None
    file_id = _drive_find_file_id(service, name, folder_id)
    if not file_id:
        return pd.DataFrame()
    raw = _drive_download_bytes(service, file_id)
    try:
        return pd.read_parquet(io.BytesIO(raw))
    except Exception:
        try:
            return pd.read_csv(io.BytesIO(raw))
        except Exception:
            return pd.DataFrame()

def save_parquet_to_drive(name: str, df: pd.DataFrame):
    service = _get_drive_service()
    folder_id = DRIVE_FOLDER_ID or None
    file_id = _drive_find_file_id(service, name, folder_id)
    buf = io.BytesIO()
    df.to_parquet(buf, index=False)
    _drive_upload_bytes(service, name, buf.getvalue(), "application/vnd.apache.parquet", folder_id, file_id)
    load_parquet_from_drive.clear()

# =========================
# DuckDB Load/Save (from/to Drive)
# =========================
def _download_duckdb_rw(name: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    try:
        service = _get_drive_service()
        if DUCKDB_FILE_ID:
            try:
                f = _drive_get_file_by_id(service, DUCKDB_FILE_ID)
                f = _drive_resolve_shortcut(service, f)
                raw = _drive_download_bytes(service, f["id"])
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"-{name}")
                tmp.write(raw); tmp.flush(); tmp.close()
                return tmp.name, f["id"], None
            except Exception as e:
                return None, None, f"DUCKDB_FILE_ID lookup failed: {e}"

        file_id = _drive_find_file_id(service, name, DRIVE_FOLDER_ID or None)
        if not file_id:
            where = DRIVE_FOLDER_ID or "all drives visible to the service account"
            return None, None, f"File '{name}' not found in {where}."
        raw = _drive_download_bytes(service, file_id)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"-{name}")
        tmp.write(raw); tmp.flush(); tmp.close()
        return tmp.name, file_id, None
    except Exception as e:
        return None, None, f"DuckDB download error: {e}"

def _upload_duckdb_back(local_path: str, file_id: str, name: str):
    service = _get_drive_service()
    with open(local_path, "rb") as f:
        data = f.read()
    _drive_upload_bytes(service, name, data, "application/octet-stream", DRIVE_FOLDER_ID or None, file_id)

@st.cache_data(ttl=300)
def load_from_duckdb(db_name: str, _schema_version: int = 2) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str,str]]:
    diags = {
        "db_file": db_name, "found": "no", "records_rows": "0", "roster_rows": "0", "note": "",
        "folder_id": DRIVE_FOLDER_ID or "(empty)", "file_id": DUCKDB_FILE_ID or "(not set)",
        "sa_email": SA_INFO.get("client_email", "<unknown>"),
    }
    local_path, file_id, err = _download_duckdb_rw(db_name)
    if err:
        diags["note"] = err
        return pd.DataFrame(), pd.DataFrame(), diags

    diags["found"] = "yes"
    try:
        con = duckdb.connect(local_path, read_only=True)
    except Exception as e:
        diags["note"] = f"connect error: {e}"
        return pd.DataFrame(), pd.DataFrame(), diags

    try:
        df_records = con.execute("SELECT Center, Language, Shift, Metric, Date, Value FROM records").df()
    except Exception:
        try:
            df_records = con.execute("SELECT Center, Language, Shift, Date, Value FROM records").df()
            df_records["Metric"] = "Requested"
            df_records = df_records[["Center","Language","Shift","Metric","Date","Value"]]
        except Exception as e:
            df_records = pd.DataFrame()
            diags["note"] += f" | records read error: {e}"

    try:
        info = con.execute("PRAGMA table_info('roster_long')").df()
        if info.empty: raise Exception("table 'roster_long' not found")
        name_map = {str(n).strip().lower(): str(n).strip() for n in info["name"].astype(str)}
        def has(col): return col.lower() in name_map
        def col(col): return name_map.get(col.lower(), col)
        select_bits = []
        def add_varchar(target_name, source_name=None):
            if source_name and has(source_name):
                select_bits.append(f"{col(source_name)} AS {target_name}")
            elif has(target_name):
                select_bits.append(f"{col(target_name)}")
            else:
                select_bits.append(f"CAST(NULL AS VARCHAR) AS {target_name}")
        add_varchar("AgentID"); add_varchar("EmpID"); add_varchar("AgentName"); add_varchar("TLName")
        add_varchar("Status"); add_varchar("WorkMode"); add_varchar("Center"); add_varchar("Location")
        add_varchar("Language"); add_varchar("SecondaryLanguage"); add_varchar("LOB"); add_varchar("FTPT")
        add_varchar("BaseShift")
        select_bits.append(f"{col('Date')}" if has("Date") else "CAST(NULL AS DATE) AS Date")
        if has("Shift"): select_bits.append(f"{col('Shift')}")
        elif has("Reason"): select_bits.append(f"{col('Reason')} AS Shift")
        else: select_bits.append("CAST(NULL AS VARCHAR) AS Shift")
        sql = "SELECT " + ", ".join(select_bits) + " FROM roster_long"
        df_roster = con.execute(sql).df()
    except Exception as e:
        df_roster = pd.DataFrame()
        diags["note"] += f" | roster_long read error: {e}"

    try: con.close()
    except Exception: pass

    diags["records_rows"] = str(0 if df_records is None or df_records.empty else len(df_records))
    diags["roster_rows"]  = str(0 if df_roster is None or df_roster.empty else len(df_roster))
    return df_records, df_roster, diags

def _ensure_duckdb_schema(con: duckdb.DuckDBPyConnection):
    # base tables
    con.execute("""
        CREATE TABLE IF NOT EXISTS records (
            Center VARCHAR, Language VARCHAR, Shift VARCHAR,
            Metric VARCHAR, Date DATE, Value DOUBLE
        );
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS roster_long (
            AgentID VARCHAR, EmpID VARCHAR, AgentName VARCHAR, TLName VARCHAR,
            Status VARCHAR, WorkMode VARCHAR, Center VARCHAR, Location VARCHAR,
            Language VARCHAR, SecondaryLanguage VARCHAR, LOB VARCHAR, FTPT VARCHAR,
            BaseShift VARCHAR, Date DATE, Shift VARCHAR
        );
    """)

    # submissions (approval flow)
    con.execute("""
        CREATE TABLE IF NOT EXISTS submissions (
            request_id VARCHAR PRIMARY KEY,
            mode VARCHAR,                 -- 'Single' | 'Bulk'
            center VARCHAR,
            uploaded_by VARCHAR,
            from_date DATE,
            to_date DATE,
            status VARCHAR,               -- 'Pending' | 'Approved' | 'Deny'
            reason VARCHAR,
            created_at TIMESTAMP,
            updated_at TIMESTAMP
        );
    """)

    con.execute("""
        CREATE TABLE IF NOT EXISTS submission_items (
            request_id VARCHAR,
            AgentID VARCHAR,
            Date DATE,
            Shift VARCHAR,
            TLName VARCHAR,
            Status VARCHAR,
            Language VARCHAR,
            LOB VARCHAR,
            WorkMode VARCHAR,
            PRIMARY KEY (request_id, AgentID, Date)
        );
    """)

    # center edit access switches
    con.execute("""
        CREATE TABLE IF NOT EXISTS center_permissions (
            center VARCHAR PRIMARY KEY,
            can_edit BOOLEAN
        );
    """)

    # admin-managed shift codes
    con.execute("""
        CREATE TABLE IF NOT EXISTS shift_codes (
            code VARCHAR PRIMARY KEY,
            label VARCHAR,
            category VARCHAR,             -- e.g., OFF, LEAVE, ACTIVITY
            is_time_code BOOLEAN,         -- if FALSE, passes w/o time validation
            min_duration_minutes INTEGER, -- optional
            notes VARCHAR,
            is_enabled BOOLEAN
        );
    """)

    # audit log
    con.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            log_id VARCHAR,               -- uuid
            action VARCHAR,
            center VARCHAR,
            actor VARCHAR,
            info VARCHAR,
            created_at TIMESTAMP
        );
    """)

def duckdb_upsert_records(local_db_path: str, file_id: str, name_in_drive: str, new_rows: pd.DataFrame) -> int:
    if new_rows.empty: return 0
    need_cols = ["Center","Language","Shift","Metric","Date","Value"]
    for c in need_cols:
        if c not in new_rows.columns:
            raise RuntimeError(f"Missing column in new Requested rows: {c}")
    new_rows = new_rows.copy()
    new_rows["Metric"] = new_rows["Metric"].fillna("Requested")
    new_rows["Date"] = pd.to_datetime(new_rows["Date"]).dt.date

    con = duckdb.connect(local_db_path, read_only=False)
    _ensure_duckdb_schema(con)
    con.register("new_records_df", new_rows)
    con.execute("""
        MERGE INTO records AS t
        USING new_records_df AS s
        ON t.Center = s.Center
           AND t.Language = s.Language
           AND t.Shift = s.Shift
           AND t.Date = s.Date
        WHEN MATCHED THEN UPDATE SET
            Metric = s.Metric, Value = s.Value
        WHEN NOT MATCHED THEN INSERT (Center, Language, Shift, Metric, Date, Value)
        VALUES (s.Center, s.Language, s.Shift, s.Metric, s.Date, s.Value);
    """)
    con.close()
    _upload_duckdb_back(local_db_path, file_id, name_in_drive)
    load_from_duckdb.clear()
    return len(new_rows)

def duckdb_replace_roster(local_db_path: str, file_id: str, name_in_drive: str, roster_df: pd.DataFrame) -> int:
    if roster_df.empty: return 0
    roster_df = roster_df.copy()
    roster_df["Date"] = pd.to_datetime(roster_df["Date"]).dt.date
    con = duckdb.connect(local_db_path, read_only=False)
    _ensure_duckdb_schema(con)
    con.register("new_roster_df", roster_df)
    con.execute("DROP TABLE IF EXISTS roster_long;")
    con.execute("CREATE TABLE roster_long AS SELECT * FROM new_roster_df;")
    inserted = con.execute("SELECT COUNT(*) FROM roster_long;").fetchone()[0]
    con.close()
    _upload_duckdb_back(local_db_path, file_id, name_in_drive)
    load_from_duckdb.clear()
    return int(inserted)

def duckdb_replace_roster_for_center(local_db_path: str, file_id: str, name_in_drive: str,
                                     roster_df: pd.DataFrame, center: str) -> int:
    if roster_df.empty:
        return 0
    df = roster_df.copy()
    if "Center" not in df.columns:
        df["Center"] = center
    else:
        df["Center"] = df["Center"].fillna(center)
    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    con = duckdb.connect(local_db_path, read_only=False)
    _ensure_duckdb_schema(con)
    con.execute("DELETE FROM roster_long WHERE Center = ?", [center])
    con.register("new_center_roster", df)
    con.execute("INSERT INTO roster_long SELECT * FROM new_center_roster")
    inserted = con.execute("SELECT COUNT(*) FROM roster_long WHERE Center = ?", [center]).fetchone()[0]
    con.close()
    _upload_duckdb_back(local_db_path, file_id, name_in_drive)
    load_from_duckdb.clear()
    return int(inserted)
import uuid
from datetime import datetime as _dt

def _audit_log(con: duckdb.DuckDBPyConnection, action: str, center: str, actor: str, info: str):
    con.execute("""
        INSERT INTO audit_log (log_id, action, center, actor, info, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, [str(uuid.uuid4()), action, center, actor, info, _dt.utcnow()])

def submit_roster_request(mode: str, center: str, uploaded_by: str,
                          from_date: date, to_date: date, items_df: pd.DataFrame) -> str:
    """Create a Pending submission + item rows; returns request_id."""
    request_id = str(uuid.uuid4())
    local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
    if err or not local_db or not file_id:
        raise RuntimeError(f"DuckDB issue: {err or 'File not found.'}")
    con = duckdb.connect(local_db, read_only=False)
    _ensure_duckdb_schema(con)
    now = _dt.utcnow()
    con.execute("""
        INSERT INTO submissions (request_id, mode, center, uploaded_by, from_date, to_date, status, reason, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, 'Pending', '', ?, ?)
    """, [request_id, mode, center, uploaded_by, from_date, to_date, now, now])

    if not items_df.empty:
        con.register("items_src", items_df)
        con.execute("""
            INSERT INTO submission_items
            SELECT ?, AgentID, Date, Shift, TLName, Status, Language, LOB, WorkMode
            FROM items_src
        """, [request_id])

    _audit_log(con, f"Submit-{mode}", center, uploaded_by, f"{len(items_df)} rows; {from_date}..{to_date}")
    con.close()
    _upload_duckdb_back(local_db, file_id, DUCKDB_FILE_NAME)
    load_from_duckdb.clear()
    return request_id

def admin_approve_request(request_id: str, actor_email: str):
    """Commit items into roster_long; mark submission Approved."""
    local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
    if err or not local_db or not file_id:
        raise RuntimeError(f"DuckDB issue: {err or 'File not found.'}")
    con = duckdb.connect(local_db, read_only=False)
    _ensure_duckdb_schema(con)

    # load header
    hdr = con.execute("SELECT center, from_date, to_date FROM submissions WHERE request_id = ?", [request_id]).df()
    if hdr.empty:
        con.close(); raise RuntimeError("Request not found")
    center = str(hdr.iloc[0]["center"])

    # apply rows (upsert per AgentID+Date)
    con.execute("CREATE TEMP TABLE _items AS SELECT * FROM submission_items WHERE request_id = ?", [request_id])
    con.execute("""
        UPDATE roster_long AS r
        SET Shift = i.Shift,
            TLName = COALESCE(NULLIF(i.TLName,''), r.TLName),
            Status = COALESCE(NULLIF(i.Status,''), r.Status),
            Language = COALESCE(NULLIF(i.Language,''), r.Language),
            LOB = COALESCE(NULLIF(i.LOB,''), r.LOB),
            WorkMode = COALESCE(NULLIF(i.WorkMode,''), r.WorkMode)
        FROM _items AS i
        WHERE r.Center = ?
          AND r.AgentID = i.AgentID
          AND r.Date = i.Date
    """, [center])

    con.execute("""
        INSERT INTO roster_long
        SELECT i.AgentID, NULL, NULL, NULL, NULL, NULL, ?, NULL, NULL, NULL, NULL, NULL, NULL, i.Date, i.Shift
        FROM _items i
        LEFT JOIN roster_long r ON r.Center = ? AND r.AgentID = i.AgentID AND r.Date = i.Date
        WHERE r.AgentID IS NULL
    """, [center, center])

    con.execute("UPDATE submissions SET status='Approved', updated_at=? WHERE request_id=?", [_dt.utcnow(), request_id])
    _audit_log(con, "Approve", center, actor_email, f"request_id={request_id}")

    con.close()
    _upload_duckdb_back(local_db, file_id, DUCKDB_FILE_NAME)
    load_from_duckdb.clear()

def admin_deny_request(request_id: str, actor_email: str, reason: str):
    local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
    if err or not local_db or not file_id:
        raise RuntimeError(f"DuckDB issue: {err or 'File not found.'}")
    con = duckdb.connect(local_db, read_only=False)
    _ensure_duckdb_schema(con)
    con.execute("UPDATE submissions SET status='Deny', reason=?, updated_at=? WHERE request_id=?",
                [reason or "(no reason)", _dt.utcnow(), request_id])
    # capture center for log
    hdr = con.execute("SELECT center FROM submissions WHERE request_id=?", [request_id]).df()
    center = str(hdr.iloc[0]["center"]) if not hdr.empty else ""
    _audit_log(con, "Deny", center, actor_email, f"request_id={request_id}; reason={reason}")
    con.close()
    _upload_duckdb_back(local_db, file_id, DUCKDB_FILE_NAME)
    load_from_duckdb.clear()
def center_can_edit(center: str) -> bool:
    if not DUCKDB_FILE_NAME:
        return True  # Parquet mode: keep current behavior
    local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
    if err or not local_db or not file_id:
        return True
    con = duckdb.connect(local_db, read_only=True)
    try:
        _ensure_duckdb_schema(con)
        df = con.execute("SELECT can_edit FROM center_permissions WHERE center = ?", [center]).df()
        if df.empty:
            return True
        return bool(df.iloc[0]["can_edit"])
    except Exception:
        return True
    finally:
        con.close()

# =========================
# Parsing helpers (Requested workbook)
# =========================
def _is_date_like(x) -> bool:
    try:
        if pd.isna(x): return False
        pd.to_datetime(x); return True
    except Exception:
        return False

def _is_language_token(x) -> bool:
    if not isinstance(x, str): return False
    s = x.strip()
    if not s or s.lower().startswith("shift"): return False
    if any(ch.isdigit() for ch in s): return False
    for ch in s:
        if not (ch.isalpha() or ch.isspace() or ch in "-_+/()"):
            return False
    return True

def _find_language_header(df: pd.DataFrame, shift_row: int, shift_col: int, win: int = 16, rows_up: int = 3):
    ncols = df.shape[1]
    for up in range(1, rows_up + 1):
        r = shift_row - up
        if r < 0: break
        for c in range(shift_col, max(-1, shift_col - win), -1):
            v = df.iat[r, c]
            if _is_language_token(v): return r, c, str(v).strip()
        for c in range(shift_col + 1, min(ncols, shift_col + 1 + win)):
            v = df.iat[r, c]
            if _is_language_token(v): return r, c, str(v).strip()
    return None, None, None

def _collect_contiguous_dates_in_row(df: pd.DataFrame, row: int, start_col: int):
    dates, cols = [], []
    c = start_col
    while c < df.shape[1] and _is_date_like(df.iat[row, c]):
        d = pd.to_datetime(df.iat[row, c]).date()
        dates.append(d); cols.append(c); c += 1
    return dates, cols

SHIFT_SPLIT = re.compile(r"\s*/\s*")
WOCL_VARIANTS = re.compile(r"^\s*W\s*O\s*[\+&/ ]\s*C\s*L\s*$", re.I)
TIME_RANGE_RE = re.compile(r"""^\s*(?:(?:\d{1,2}[:.]?\d{2})\s*-\s*(?:\d{1,2}[:.]?\d{2}))
                                (?:\s*/\s*(?:\d{1,2}[:.]?\d{2})\s*-\s*(?:\d{1,2}[:.]?\d{2}))*\s*$""", re.X)

def _is_off_code(s: str) -> Optional[str]:
    t = re.sub(r"\s+", "", str(s)).upper()
    if t in {"WO","OFF"}: return "WO"
    if t in {"CL","CO","COMPOFF","COMP-OFF"}: return "CL"
    if WOCL_VARIANTS.fullmatch(str(s)): return "WO+CL"
    return None

def _is_valid_shift_label(s: str) -> bool:
    if _is_off_code(s): return True
    return bool(TIME_RANGE_RE.fullmatch(str(s)))

def _normalize_shift_label(s: str) -> Optional[str]:
    off = _is_off_code(s)
    if off: return off
    txt = str(s).strip()
    if not _is_valid_shift_label(txt): return None
    parts = SHIFT_SPLIT.split(txt)
    norm_parts = []
    for part in parts:
        part = part.strip().upper().replace("HRS","").replace("HR","").replace(".",":")
        m = re.fullmatch(r"(\d{1,2})[:]?(\d{2})\s*-\s*(\d{1,2})[:]?(\d{2})", part)
        if not m: return None
        sh, sm, eh, em = map(int, m.groups())
        sh = max(0, min(23, sh)); eh = max(0, min(23, eh))
        sm = max(0, min(59, sm)); em = max(0, min(59, em))
        norm_parts.append(f"{sh:02d}{sm:02d}-{eh:02d}{em:02d}")
    return "/".join(norm_parts)
# ===== SOP Allowed Values & Limits =====
ALLOWED_STATUS = {
    "BAU","Nesting Week 4","Nesting Week 3","Nesting Week 2","Nesting Week 1",
    "AUX Trainer","AUX QA","Floor Support","OJT","Aux TL","Certification","Training"
}
ALLOWED_PRIMARY_LANGUAGE = {
    "English","Kannada","Telugu","Tamil","Bengali","Hindi","Hindi_GIG","Bengali_GIG",
    "English_GIG","Telugu_GIG","Malayalam"
}
ALLOWED_WORKMODE = {"WFO","WFH"}
MAX_EDIT_WINDOW_DAYS = 7  # ≤7 days per SOP

def parse_center_sheet(xls: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(xls, sheet_name, header=None)
    if df.empty:
        return pd.DataFrame(columns=['Center','Language','Shift','Metric','Date','Value'])
    nrows, ncols = df.shape
    out = []
    for i in range(1, nrows):
        for j in range(ncols):
            cell = df.iat[i, j]
            if isinstance(cell, str) and re.search(r"\bshift(?:\s*name|s)?\b", cell, flags=re.IGNORECASE):
                _, _, lang = _find_language_header(df, i, j)
                if not lang: continue
                first_date_col = None
                for c in range(j + 1, ncols):
                    if _is_date_like(df.iat[i, c]): first_date_col = c; break
                if first_date_col is None: continue
                dates, date_cols = _collect_contiguous_dates_in_row(df, i, first_date_col)
                if not dates: continue
                r, blanks = i + 1, 0
                while r < nrows:
                    lbl = df.iat[r, j]
                    if pd.isna(lbl) or str(lbl).strip() == "":
                        blanks += 1
                        if blanks >= 2: break
                        r += 1; continue
                    blanks = 0
                    shift_label_norm = _normalize_shift_label(str(lbl).strip())
                    if not shift_label_norm:
                        r += 1; continue
                    for d, dc in zip(dates, date_cols):
                        v = pd.to_numeric(df.iat[r, dc], errors="coerce")
                        if pd.notna(v):
                            out.append((sheet_name, lang, shift_label_norm, "Requested", d, float(v)))
                    r += 1
    cols = ['Center','Language','Shift','Metric','Date','Value']
    return pd.DataFrame(out, columns=cols) if out else pd.DataFrame(columns=cols)

def parse_workbook_to_df(file_bytes: bytes) -> pd.DataFrame:
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    EXCLUDED = {"Settings","Roster","Overall"}
    centers = [s for s in xls.sheet_names if s.casefold() not in {x.casefold() for x in EXCLUDED}]
    parts = []
    for s in centers:
        try:
            d = parse_center_sheet(xls, s)
            if not d.empty: parts.append(d)
        except Exception as e:
            st.warning(f"Failed parsing '{s}': {e}")
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=['Center','Language','Shift','Metric','Date','Value'])

def read_roster_sheet(file_bytes: bytes) -> pd.DataFrame:
    try:
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name="Roster")
    except Exception as e:
        st.warning(f"Could not read 'Roster' sheet: {e}")
        return pd.DataFrame()

# =========================
# Data model selection (prefer DuckDB)
# =========================
RECORDS_FILE = "records.parquet"
ROSTER_FILE  = "roster_long.parquet"

def _centers_union(records: pd.DataFrame, roster: pd.DataFrame) -> List[str]:
    a = set(records["Center"].unique().tolist()) if not records.empty and "Center" in records else set()
    b = set(roster["Center"].unique().tolist())  if not roster.empty  and "Center" in roster  else set()
    cs = sorted([c for c in (a | b) if c])
    return ["Overall"] + cs

def _languages_union(records: pd.DataFrame, roster: pd.DataFrame, center: Optional[str]) -> List[str]:
    def unique_lang(df):
        return set(df["Language"].dropna().astype(str).str.strip().unique().tolist()) if not df.empty and "Language" in df else set()
    if center in (None, "", "Overall"):
        return sorted([l for l in (unique_lang(records) | unique_lang(roster)) if l])
    else:
        a = unique_lang(records[records["Center"]==center]) if "Center" in records else set()
        b = unique_lang(roster[roster["Center"]==center]) if "Center" in roster else set()
        return sorted([l for l in (a | b) if l])
def _ensure_roster_schema(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensures roster has canonical columns: 'Date' and 'Shift'.
    Accepts common variants like 'date', 'DATE', or 'Reason' (for Shift).
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=["Date", "Shift"])
    out = df.copy()

    # Map variants -> canonical
    cols_lower = {c.lower(): c for c in out.columns}

    def _alias(*names):
        for n in names:
            if n.lower() in cols_lower:
                return cols_lower[n.lower()]
        return None

    date_col  = _alias("Date", "date", "DATE")
    shift_col = _alias("Shift", "shift", "SHIFT", "Reason", "reason")

    if date_col and date_col != "Date":
        out = out.rename(columns={date_col: "Date"})
    if shift_col and shift_col != "Shift":
        out = out.rename(columns={shift_col: "Shift"})

    return out

def load_store() -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str,str]]:
    diags = {}
    if DUCKDB_FILE_NAME:
        try:
            records, roster, d = load_from_duckdb(DUCKDB_FILE_NAME, _schema_version=2)
        except Exception:
            try: load_from_duckdb.clear()
            except Exception: pass
            records, roster, d = load_from_duckdb(DUCKDB_FILE_NAME, _schema_version=2)
        diags.update({"source": "duckdb", **d})
    else:
        records = load_parquet_from_drive(RECORDS_FILE)
        roster  = load_parquet_from_drive(ROSTER_FILE)
        diags.update({"source":"parquet",
                      "records_rows":str(0 if records.empty else len(records)),
                      "roster_rows": str(0 if roster.empty else len(roster)),
                      "found":"n/a","db_file":"n/a","note":"",
                      "folder_id": DRIVE_FOLDER_ID or "(empty)",
                      "file_id": DUCKDB_FILE_ID or "(not set)",
                      "sa_email": SA_INFO.get("client_email","<unknown>")})

    if not records.empty:
        records["Date"] = pd.to_datetime(records["Date"]).dt.date
        if "Metric" not in records or records["Metric"].isna().all():
            records["Metric"] = "Requested"
        if (records["Metric"] == "Requested").any():
            records = records[records["Metric"] == "Requested"].copy()

    if not roster.empty:
        # Ensure canonical columns exist
        roster = _ensure_roster_schema(roster)

        if "Date" in roster.columns:
            roster["Date"] = pd.to_datetime(roster["Date"], errors="coerce").dt.date

        # Clean important columns if present
        for c in ["AgentID", "Language", "Center", "LOB", "Shift"]:
            if c in roster.columns:
                roster[c] = roster[c].astype(str).str.strip()

        if "Shift" in roster.columns:
            roster["Shift"] = roster["Shift"].apply(
                lambda s: (_normalize_shift_label(s) or str(s).strip().upper())
            )

    return records, roster, diags

# =========================
# Aggregation + Views
# =========================
def _pretty_int(x: float) -> int:
    try: return int(round(float(x)))
    except Exception: return 0

_TIME_RE_SINGLE = re.compile(r"^\d{4}-\d{4}$")

def _minutes_of(label: str) -> Optional[int]:
    s = str(label).strip()
    if not _TIME_RE_SINGLE.fullmatch(s): return None
    sh, sm, eh, em = int(s[0:2]), int(s[2:4]), int(s[5:7]), int(s[7:9])
    start = sh*60 + sm; end = eh*60 + em
    return (end - start) if end > start else None
def _duration_minutes(label: str) -> Optional[int]:
    """Return duration minutes for a single-segment 'HHMM-HHMM' (no splits)."""
    m = re.fullmatch(r"^\s*(\d{2})(\d{2})-(\d{2})(\d{2})\s*$", str(label))
    if not m: return None
    sh, sm, eh, em = map(int, m.groups())
    start = sh*60 + sm
    end   = eh*60 + em
    if end <= start:   # disallow negative or cross-midnight for validation calc
        return None
    return end - start

def _is_dynamic_code(s: str) -> bool:
    s = str(s).strip().upper()
    return s in {"WO","CL","FS"}  # extended via shift_codes table as well

def validate_shift_value(raw: str, con: duckdb.DuckDBPyConnection | None = None) -> tuple[bool,str]:
    """
    Returns (is_valid, reason).
    Rules:
      - Dynamic codes: WO, CL, FS (plus enabled shift_codes with is_time_code=False)
      - Single time range: HHMM-HHMM with duration >= 540 minutes (9h)
      - Split ranges: HHMM-HHMM/HHMM-HHMM must be exactly 5h + 4h (order agnostic)
    """
    s = (raw or "").strip().upper()
    if not s:
        return False, "Empty shift"

    # admin-managed codes
    if con is not None:
        try:
            cfg = con.execute("SELECT code, is_time_code FROM shift_codes WHERE is_enabled = TRUE").df()
            codes = {str(r["code"]).upper(): bool(r["is_time_code"]) for _, r in cfg.iterrows()} if not cfg.empty else {}
        except Exception:
            codes = {}
    else:
        codes = {}

    # dynamic/non-time codes
    if _is_dynamic_code(s) or (s in codes and codes[s] is False):
        return True, ""

    # time patterns
    parts = [p.strip() for p in s.split("/") if p.strip()]
    if len(parts) == 1:
        d = _duration_minutes(parts[0])
        if d is None:
            return False, "Invalid time pattern for single range"
        if d < 540:
            return False, "Single range must be ≥ 9h"
        return True, ""
    elif len(parts) == 2:
        d1 = _duration_minutes(parts[0])
        d2 = _duration_minutes(parts[1])
        if d1 is None or d2 is None:
            return False, "Invalid time pattern in split"
        if sorted([d1, d2]) != [240, 300]:
            return False, "Split must be 5h + 4h"
        return True, ""
    else:
        return False, "Too many segments (max 2)"
def validate_allowed_values(row: dict) -> list[str]:
    """Return list of error messages for a roster edit row."""
    errs = []
    status = str(row.get("Status","")).strip()
    lang   = str(row.get("Language","")).strip()
    wm     = str(row.get("WorkMode","")).strip()

    if status and status not in ALLOWED_STATUS:
        errs.append(f"Invalid Status: {status}")
    if lang and lang not in ALLOWED_PRIMARY_LANGUAGE:
        errs.append(f"Invalid Primary Language: {lang}")
    if wm and wm not in ALLOWED_WORKMODE:
        errs.append(f"Invalid WorkMode: {wm}")
    return errs

def validate_window(from_date: date, to_date: date) -> Optional[str]:
    if (to_date - from_date).days > (MAX_EDIT_WINDOW_DAYS - 1):
        return f"Window must be ≤ {MAX_EDIT_WINDOW_DAYS} days"
    return None

@st.cache_data(show_spinner=False, ttl=600)
def dedup_requested_split_pairs(pivot_df: pd.DataFrame) -> pd.DataFrame:
    if pivot_df.empty: return pivot_df
    df = pivot_df.copy()
    idx = list(df.index)
    single_seg_minutes = {s: (None if "/" in str(s) else _minutes_of(str(s))) for s in idx}
    for col in df.columns:
        val_to_5h, val_to_4h = {}, {}
        col_values = df[col]
        for s in idx:
            mins = single_seg_minutes.get(s)
            if mins not in (300, 240): continue
            v = col_values.get(s)
            if pd.isna(v): continue
            try: vv = float(v)
            except Exception: continue
            (val_to_5h if mins==300 else val_to_4h).setdefault(vv, set()).add(s)
        for v in set(val_to_4h.keys()) & set(val_to_5h.keys()):
            for row_4h in val_to_4h[v]:
                df.at[row_4h, col] = 0.0
    return df

@st.cache_data(show_spinner=False, ttl=600)
def transform_to_interval_view(shift_df: pd.DataFrame) -> pd.DataFrame:
    if shift_df.empty: return pd.DataFrame(index=range(24))
    cols_dates = [pd.to_datetime(c).date() for c in shift_df.columns]
    shift_df = shift_df.copy(); shift_df.columns = cols_dates
    interval_df = pd.DataFrame(0.0, index=range(24), columns=cols_dates)
    def _minutes_pair(text: str) -> Optional[tuple[int,int]]:
        t = str(text).upper().strip().replace("HRS","").replace("HR","").replace(".",":")
        m = re.fullmatch(r"\s*(\d{1,2})[:]?(\d{2})\s*-\s*(\d{1,2})[:]?(\d{2})\s*", t)
        if not m: return None
        sh, sm, eh, em = map(int, m.groups())
        if not (0<=sh<=23 and 0<=eh<=23 and 0<=sm<=59 and 0<=em<=59): return None
        return sh*60+sm, eh*60+em
    def _add_span(to_date, start_min, end_min, counts):
        if end_min <= start_min: return
        if to_date not in interval_df.columns: interval_df[to_date] = 0.0
        for h in range(24):
            dur = max(0, min(end_min, (h+1)*60) - max(start_min, h*60))
            if dur > 0:
                interval_df.at[h, to_date] += float(counts) * (dur / 60.0)
    for shift_label, counts_series in shift_df.iterrows():
        if str(shift_label).upper() in {"WO","CL","WO+CL"}: continue
        for part in str(shift_label).split("/"):
            part = part.strip()
            if re.fullmatch(r"\d{4}-\d{4}", part):
                smin = int(part[:2]) * 60 + int(part[2:4])
                emin = int(part[5:7]) * 60 + int(part[7:9])
            else:
                pair = _minutes_pair(part)
                if not pair: continue
                smin, emin = pair
            for day, count in counts_series.items():
                if pd.isna(count) or float(count) == 0.0: continue
                if emin > smin:
                    _add_span(day, smin, emin, float(count))
                else:
                    _add_span(day, smin, 24*60, float(count))
                    _add_span(day + timedelta(days=1), 0, emin, float(count))
    final_cols = sorted(set(cols_dates) | set(interval_df.columns))
    return interval_df.reindex(columns=final_cols, fill_value=0.0)

# =========================
# Query utilities
# =========================
def union_dates(records: pd.DataFrame, roster: pd.DataFrame, center: Optional[str]) -> List[date]:
    rs = set()
    if not records.empty:
        rs |= set(records["Date"].unique().tolist()) if center in (None, "", "Overall") else set(records.loc[records["Center"]==center, "Date"].unique().tolist())
    if not roster.empty:
        rs |= set(roster["Date"].unique().tolist()) if center in (None, "", "Overall") else set(roster.loc[roster["Center"]==center, "Date"].unique().tolist())
    return sorted(list(rs))

@st.cache_data(show_spinner=False, ttl=300)
def pivot_requested(records, center, langs, start, end) -> pd.DataFrame:
    if records.empty or not langs: return pd.DataFrame()
    df = records[(records["Date"].between(start, end)) & (records["Language"].isin(langs))].copy()
    if center != "Overall": df = df[df["Center"] == center]
    if df.empty: return pd.DataFrame()
    p = df.groupby(["Shift","Date"], as_index=False)["Value"].sum()
    p = p.pivot(index="Shift", columns="Date", values="Value").fillna(0.0)
    p = p.reindex(sorted(p.columns), axis=1)
    p.index.name = "Shift"
    return p

def roster_lobs(roster: pd.DataFrame, center: Optional[str], start: date, end: date) -> List[str]:
    if roster.empty or "LOB" not in roster.columns: return []
    df = roster[roster["Date"].between(start, end)]
    if center and center != "Overall": df = df[df["Center"] == center]
    return sorted([x for x in df["LOB"].dropna().astype(str).str.strip().unique().tolist() if x])

def roster_languages_for_lob(roster: pd.DataFrame, center: Optional[str], lob: Optional[str], start: date, end: date) -> List[str]:
    if roster.empty: return []
    df = roster[roster["Date"].between(start, end)]
    if center and center != "Overall": df = df[df["Center"] == center]
    if lob: df = df[df["LOB"] == lob]
    return sorted(df["Language"].dropna().astype(str).str.strip().unique().tolist())

def roster_distinct_shifts(roster: pd.DataFrame, center: Optional[str], start: date, end: date,
                           lob: Optional[str], language: Optional[str]) -> List[str]:
    if roster.empty: return []
    df = roster[roster["Date"].between(start, end)]
    if center and center != "Overall": df = df[df["Center"] == center]
    if lob: df = df[df["LOB"] == lob]
    if language: df = df[df["Language"] == language]
    pat = re.compile(r"^\d{4}-\d{4}(?:/\d{4}-\d{4})*$")
    vals = []
    for s in df["Shift"].dropna().astype(str):
        t = s.strip()
        if pat.fullmatch(t): vals.append(t)
    combined = [s for s in vals if "/" in s]
    parts_set = set()
    for s in combined:
        parts_set.update(p.strip() for p in s.split("/") if p.strip())
    vals_clean = [s for s in vals if ("/" in s) or (s not in parts_set)]
    def _key(s):
        first = s.split("/")[0]
        return (int(first[:2]) * 60 + int(first[2:4]), s)
    return sorted(dict.fromkeys(vals_clean), key=_key)

def roster_nonshift_codes(roster: pd.DataFrame, center: Optional[str], start: date, end: date,
                          lob: Optional[str], language: Optional[str]) -> List[str]:
    if roster.empty: return []
    df = roster[roster["Date"].between(start, end)]
    if center and center != "Overall": df = df[df["Center"] == center]
    if lob: df = df[df["LOB"] == lob]
    if language: df = df[df["Language"] == language]
    pat = re.compile(r"^\d{4}-\d{4}(?:/\d{4}-\d{4})*$")
    vals = []
    for s in df["Shift"].dropna().astype(str):
        t = s.strip()
        if t and not pat.fullmatch(t):
            vals.append(t)
    return sorted(dict.fromkeys(vals), key=lambda x: x.upper())

@st.cache_data(show_spinner=False, ttl=300)
def pivot_roster_counts(roster: pd.DataFrame,
                        center: Optional[str],
                        languages_: List[str],
                        start: date,
                        end: date) -> pd.DataFrame:
    if roster.empty or not languages_:
        return pd.DataFrame()
    df = roster[roster["Date"].between(start, end)].copy()
    if center and center != "Overall":
        df = df[df["Center"] == center]
    df = df[df["Language"].isin(languages_)]
    if df.empty:
        return pd.DataFrame()
    time_rows: list[tuple[date, str]] = []
    code_rows: list[tuple[date, str]] = []
    for _, row in df.iterrows():
        dt = row["Date"]
        raw = str(row.get("Shift", "")).strip()
        norm = _normalize_shift_label(raw)
        if not norm:
            continue
        if norm in {"WO", "CL", "WO+CL"}:
            code_rows.append((dt, "WO" if norm == "WO" else ("CL" if norm == "CL" else "WO+CL")))
            continue
        parts = [p.strip() for p in str(norm).split("/") if p.strip()]
        for p in parts:
            if re.fullmatch(r"\d{4}-\d{4}", p):
                time_rows.append((dt, p))
    p = pd.DataFrame()
    if time_rows:
        tdf = pd.DataFrame(time_rows, columns=["Date", "Shift"])
        p = tdf.groupby(["Shift", "Date"]).size().unstack("Date", fill_value=0).astype(float)
        p = p.reindex(sorted(p.columns), axis=1)
    if code_rows:
        cdf = pd.DataFrame(code_rows, columns=["Date", "Code"])
        codes = cdf.groupby(["Date", "Code"]).size().unstack("Code", fill_value=0)
        w = codes.get("WO", 0); c = codes.get("CL", 0); wc = codes.get("WO+CL", 0)
        wocl = (w + c + wc).astype(float)
        all_cols = sorted(set(p.columns) | set(wocl.index.tolist()))
        p = (pd.DataFrame(columns=all_cols) if p.empty else p.reindex(columns=all_cols, fill_value=0.0))
        p.loc["WO+CL"] = [float(wocl.get(col, 0.0)) for col in p.columns]
    p.index.name = "Shift"
    return p

def add_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    out = df.copy()
    out.loc["Total"] = out.sum(axis=0)
    return out

def roster_agents_wide(roster: pd.DataFrame, center: Optional[str], lob: Optional[str], language: Optional[str],
                       agent_ids: Optional[List[str]], start: date, end: date) -> pd.DataFrame:
    if roster.empty: return pd.DataFrame()
    df = roster[roster["Date"].between(start, end)].copy()
    if center and center != "Overall": df = df[df["Center"] == center]
    if lob: df = df[df["LOB"] == lob]
    if language: df = df[df["Language"] == language]
    if agent_ids:
        ids = [str(x).strip() for x in agent_ids if str(x).strip()]
        df = df[df["AgentID"].isin(ids)]
    if df.empty: return pd.DataFrame()
    static_cols = [c for c in ["AgentID","AgentName","TLName","Status","WorkMode","Center","Location",
                               "Language","SecondaryLanguage","LOB","FTPT"] if c in df.columns]
    p = df.pivot_table(index=static_cols, columns="Date", values="Shift", aggfunc="first")
    if len(p.columns): p = p.reindex(sorted(p.columns), axis=1)
    p = p.reset_index()
    return p

# =========================
# Export utilities
# =========================
def export_excel(view_type: str, center: str, languages_sel: List[str], start: date, end: date,
                 records: pd.DataFrame, roster: pd.DataFrame) -> bytes:
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        req_s = dedup_requested_split_pairs(pivot_requested(records, center, languages_sel, start, end))
        ros_s = pivot_roster_counts(roster, center, languages_sel, start, end)
        all_dates = sorted(list(set(req_s.columns) | set(ros_s.columns)))
        all_shifts = sorted(list(set(req_s.index) | set(ros_s.index)))
        req_s = req_s.reindex(index=all_shifts, columns=all_dates, fill_value=0.0)
        ros_s = ros_s.reindex(index=all_shifts, columns=all_dates, fill_value=0.0)
        delt_s = ros_s - req_s
        req_s, ros_s, delt_s = add_total_row(req_s), add_total_row(ros_s), add_total_row(delt_s)

        if view_type in ["Shift_Wise Delta View", "Overall_Delta View"]:
            sheet_name = "Shift_Wise_View"
            req_s.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)
            ros_s.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1, startcol=len(req_s.columns) + 2)
            delt_s.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1,
                            startcol=len(req_s.columns) + len(ros_s.columns) + 4)
            ws = writer.sheets[sheet_name]
            for col_cells in ws.columns:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 36)

        if view_type in ["Interval_Wise Delta View", "Overall_Delta View"]:
            req_i = transform_to_interval_view(req_s.drop('Total', errors='ignore'))
            ros_i = transform_to_interval_view(ros_s.drop('Total', errors='ignore'))
            all_cols_i = sorted(list(set(req_i.columns) | set(ros_i.columns)))
            req_i = req_i.reindex(columns=all_cols_i, fill_value=0.0)
            ros_i = ros_i.reindex(columns=all_cols_i, fill_value=0.0)
            delt_i = ros_i - req_i
            req_i, ros_i, delt_i = add_total_row(req_i), add_total_row(ros_i), add_total_row(delt_i)
            sheet_name = "Interval_Wise_View"
            req_i.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1)
            ros_i.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1, startcol=len(req_i.columns) + 2)
            delt_i.to_excel(writer, sheet_name=sheet_name, index=True, startrow=1,
                            startcol=len(req_i.columns) + len(ros_i.columns) + 4)
            ws = writer.sheets[sheet_name]
            for col_cells in ws.columns:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 36)
    return output.getvalue()

# === Requested editor helpers (desktop-style) ===
def _validate_shift_label_for_edit(s: str) -> Optional[str]:
    if pd.isna(s) or str(s).strip() == "":
        return None
    return _normalize_shift_label(str(s).strip())

def _pivot_requested_one_lang(records: pd.DataFrame, center: str, language: str, start: date, end: date) -> pd.DataFrame:
    if records.empty:
        return pd.DataFrame()
    df = records[(records["Date"].between(start, end)) & (records["Language"] == language)].copy()
    if center and center != "Overall":
        df = df[df["Center"] == center]
    if df.empty:
        return pd.DataFrame()
    p = df.groupby(["Shift","Date"], as_index=False)["Value"].sum()
    p = p.pivot(index="Shift", columns="Date", values="Value").fillna(0.0)
    p = p.reindex(sorted(p.columns), axis=1)
    p.index.name = "Shift"
    return p

def _save_requested_edits(center: str, language: str, edited_df: pd.DataFrame, start: date, end: date) -> Tuple[int, str]:
    if edited_df is None or edited_df.empty:
        return 0, "Nothing to save."
    new_df = edited_df.copy()
    try:
        new_df.columns = [pd.to_datetime(c).date() for c in new_df.columns]
    except Exception:
        new_df = new_df[[c for c in new_df.columns if isinstance(c, (pd.Timestamp, date))]]
    if new_df.empty:
        return 0, "No valid date columns after validation."
    new_df = new_df.reset_index().rename(columns={new_df.index.name or "index": "Shift"})
    new_df["ShiftNorm"] = new_df["Shift"].apply(_validate_shift_label_for_edit)
    new_df = new_df[~new_df["ShiftNorm"].isna()].copy()
    if new_df.empty:
        return 0, "All rows had invalid/empty shift labels."
    new_df = new_df.drop(columns=["Shift"]).rename(columns={"ShiftNorm": "Shift"}).set_index("Shift")
    out_long = new_df.reset_index().melt(id_vars="Shift", var_name="Date", value_name="Value")
    out_long = out_long.dropna(subset=["Value"])
    if out_long.empty:
        return 0, "No numeric values to save."
    out_long["Value"] = pd.to_numeric(out_long["Value"], errors="coerce")
    out_long = out_long.dropna(subset=["Value"])
    if out_long.empty:
        return 0, "No numeric values to save."
    out_long["Center"] = center
    out_long["Language"] = language
    out_long["Metric"] = "Requested"
    out_long = out_long[["Center","Language","Shift","Metric","Date","Value"]].copy()
    if DUCKDB_FILE_NAME:
        local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
        if err or not local_db or not file_id:
            return 0, f"DuckDB issue: {err or 'File not found.'}"
        try:
            n = duckdb_upsert_records(local_db, file_id, DUCKDB_FILE_NAME, out_long)
            return n, f"Saved {n} Requested cells for {language}."
        except Exception as e:
            return 0, f"DuckDB upsert failed: {e}"
    else:
        cur = load_parquet_from_drive(RECORDS_FILE)
        if cur.empty:
            save_parquet_to_drive(RECORDS_FILE, out_long)
            load_parquet_from_drive.clear()
            return len(out_long), f"Saved {len(out_long)} Requested cells for {language}."
        keys = ["Center","Language","Shift","Date"]
        left = cur.merge(out_long[keys].drop_duplicates(), on=keys, how="left", indicator=True)
        keep = left["_merge"] == "left_only"
        base = cur[keep].copy()
        new_all = pd.concat([base, out_long], ignore_index=True)
        save_parquet_to_drive(RECORDS_FILE, new_all)
        return len(out_long), f"Saved {len(out_long)} Requested cells for {language}."

# =========================
# UI
# =========================
st.title("📊 Center Shiftwise Web Dashboard")
st.caption("Google Drive-backed • Streamlit • DuckDB/Parquet (Shared Drive compatible)")

# Load current store
# Load current store (overlay for UX)
with overlay("Loading data from Drive / DuckDB…"):
    records, roster, diags = load_store()


with st.sidebar:
    # Account + logout
    st.subheader("👤 Account")
    if user:
        st.write(f"**{user.get('name','')}**")
        st.caption(user.get('email',''))
        st.caption(f"Role: {user.get('role','viewer')}")
        if st.button("Logout", use_container_width=True):
            st.session_state.pop("user", None)
            st.rerun()
    st.divider()

    st.subheader("🔐 Google Drive Setup")
    st.write("- Service account JSON in **st.secrets['gcp_service_account']**"
             "\n- **DRIVE_FOLDER_ID** points to the Shared Drive folder with your DB"
             "\n- Prefer setting **DUCKDB_FILE_ID** (faster sync)")

    if st.button("↻ Sync from Drive", use_container_width=True):
        with overlay("Refreshing cached data…"):
            load_parquet_from_drive.clear()
            load_from_duckdb.clear()
            _load_acl_from_drive_cached.clear()
        st.rerun()

    st.divider()
    st.subheader("🧪 Data In Database ")
    st.write(f"**Records_Data:** {diags.get('records_rows')}")
    st.write(f"**Roster_Data:** {diags.get('roster_rows')}")
    if diags.get("note"): st.caption(f"_note_: {diags.get('note')}")

    if (records.empty) and (roster.empty):
        st.warning("Both `records` and `roster_long` are empty. Check filename, file/folder access, and table schemas.")

    st.divider()
    st.subheader("⬆️ Import / Fetch Files (Excel)")

    # ---- Requested: upload or fetch from Drive folder ----
    st.markdown("**Requested workbook (.xlsx)**")
    req_src = st.radio("Source", ["Upload", "Drive folder"], horizontal=True, key="req_src")
    req_file = None
    if req_src == "Upload":
        req_file = st.file_uploader("Upload Requested workbook", type=["xlsx"], key="req_up")
    else:
        if not REQUESTED_FOLDER_ID:
            st.info("Set REQUESTED_FOLDER_ID in secrets to fetch from Drive.")
        else:
            try:
                service = _get_drive_service()
                files = _drive_list_folder(REQUESTED_FOLDER_ID)
                names = [f['name'] for f in files if f.get('mimeType') != 'application/vnd.google-apps.folder']
                sel = st.selectbox("Pick a file from Drive", names)
                if sel:
                    fid = next((f['id'] for f in files if f['name']==sel), None)
                    if fid and st.button("Fetch Requested from Drive"):
                        raw = _drive_download_bytes(service, fid)
                        st.session_state["__fetched_req_bytes__"] = raw
                        st.success("Requested file fetched.")
                if "__fetched_req_bytes__" in st.session_state:
                    req_file = io.BytesIO(st.session_state["__fetched_req_bytes__"])
            except Exception as e:
                st.error(f"Drive browse error: {e}")

    if st.button("Import Requested", type="primary", use_container_width=True, disabled=(req_file is None)):
        try:
            with overlay("Parsing & upserting Requested…"):
                raw_bytes = req_file.read() if hasattr(req_file, "read") else req_file
                new_req = parse_workbook_to_df(raw_bytes)
                if new_req.empty:
                    st.warning("No Requested data parsed.")
                else:
                    if DUCKDB_FILE_NAME:
                        local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
                        if err or not local_db or not file_id:
                            st.error(f"DuckDB issue: {err or 'File not found.'}")
                        else:
                            affected = duckdb_upsert_records(local_db, file_id, DUCKDB_FILE_NAME, new_req)
                            st.success(f"Upserted Requested rows: {affected}")
                    else:
                        cur = load_parquet_from_drive(RECORDS_FILE)
                        if cur.empty:
                            out = new_req
                        else:
                            keys = ["Center", "Language", "Shift", "Date"]
                            left = cur.merge(new_req[keys].drop_duplicates(), on=keys, how="left", indicator=True)
                            keep = left["_merge"] == "left_only"
                            base = cur[keep].copy()
                            out = pd.concat([base, new_req], ignore_index=True)
                        save_parquet_to_drive(RECORDS_FILE, out)
                        st.success(f"Imported Requested rows: {len(new_req)}")
        except Exception as e:
            st.error(f"Import Requested failed: {e}")

    st.markdown("---")

    # ---- Roster: upload or fetch from Drive folder ----
    st.markdown("**Roster workbook (.xlsx)**")
    rost_src = st.radio("Source", ["Upload", "Drive folder"], horizontal=True, key="rost_src")
    rost_file = None
    if rost_src == "Upload":
        rost_file = st.file_uploader("Upload Roster workbook", type=["xlsx"], key="rost_up")
    else:
        if not ROSTER_FOLDER_ID:
            st.info("Set ROSTER_FOLDER_ID in secrets to fetch from Drive.")
        else:
            try:
                service = _get_drive_service()
                files = _drive_list_folder(ROSTER_FOLDER_ID)
                names = [f['name'] for f in files if f.get('mimeType') != 'application/vnd.google-apps.folder']
                sel = st.selectbox("Pick a roster file from Drive", names)
                if sel:
                    fid = next((f['id'] for f in files if f['name']==sel), None)
                    if fid and st.button("Fetch Roster from Drive"):
                        raw = _drive_download_bytes(service, fid)
                        st.session_state["__fetched_rost_bytes__"] = raw
                        st.success("Roster file fetched.")
                if "__fetched_rost_bytes__" in st.session_state:
                    rost_file = io.BytesIO(st.session_state["__fetched_rost_bytes__"])
            except Exception as e:
                st.error(f"Drive browse error: {e}")

    if st.button("Import Roster (replace ALL)", use_container_width=True, disabled=(rost_file is None)):
        try:
            with overlay("Importing roster & replacing table…"):
                raw_df = read_roster_sheet(rost_file.read() if hasattr(rost_file, "read") else rost_file)
                if raw_df.empty:
                    st.warning("No Roster data found (need a 'Roster' sheet).")
                else:
                    if DUCKDB_FILE_NAME:
                        local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
                        if err or not local_db or not file_id:
                            st.error(f"DuckDB issue: {err or 'File not found.'}")
                        else:
                            if "Date" not in raw_df or "Shift" not in raw_df:
                                st.error("Uploaded roster must have at least 'Date' and 'Shift' columns.")
                            else:
                                raw_df["Date"] = pd.to_datetime(raw_df["Date"]).dt.date
                                inserted = duckdb_replace_roster(local_db, file_id, DUCKDB_FILE_NAME, raw_df)
                                st.success(f"Roster replaced in DuckDB. Rows inserted: {inserted}")
                    else:
                        save_parquet_to_drive(ROSTER_FILE, raw_df)
                        st.success(f"Roster imported to Parquet. Rows: {len(raw_df)}")
        except Exception as e:
            st.error(f"Import Roster failed: {e}")

    st.divider()

# ========== Top filter row (RBAC-aware) ==========
c1, c2, c3, c4 = st.columns([1.5, 1.2, 1.3, 1.7])
with c1:
    center_vals_all = _centers_union(records, roster)  # ["Overall", ...centers]
    centers_no_overall = [c for c in center_vals_all if c != "Overall"]
    ACL = _load_acl_from_drive_cached(DRIVE_FOLDER_ID or None)
    allowed_centers, has_full_access = resolve_allowed_centers_from_email(user.get("email",""), centers_no_overall, ACL)
    is_admin = (user.get("role","viewer") == "admin")
    if has_full_access:
        center_options = center_vals_all
    else:
        center_options = allowed_centers[:]  # no "Overall" for restricted users
    if not center_options:
        st.error("Your account has no center access. Contact an admin.")
        center = None
    else:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        center = st.selectbox("Center", center_options, index=0)
with c2:
    ds = union_dates(records, roster, center)
    if ds: start_default, end_default = ds[0], ds[-1]
    else:  start_default = end_default = date.today()
    date_range = st.date_input("Date range", value=(start_default, end_default))
    start, end = (date_range if isinstance(date_range, tuple) else (start_default, end_default))
with c3:
    view_type = st.selectbox(
        "View",
        ["Shift_Wise Delta View","Interval_Wise Delta View","Overall_Delta View"] if (center is not None) else [],
        index=0 if center is not None else None
    )
with c4:
    lang_choices = _languages_union(records, roster, center) if center else []
    langs_sel = st.multiselect("Languages", lang_choices, default=(lang_choices if lang_choices else []))
    st.markdown('</div>', unsafe_allow_html=True)

can_edit_this_center = bool(center) and user_can_edit_center(
    user.get("email",""), user.get("role","viewer"), center,
    allowed_centers, has_full_access
)

# Build data for views
req_shift = dedup_requested_split_pairs(pivot_requested(records, center, langs_sel, start, end))
ros_shift = pivot_roster_counts(roster, center, langs_sel, start, end)
all_dates = sorted(list(set(req_shift.columns) | set(ros_shift.columns)))
all_shifts = sorted(list(set(req_shift.index) | set(ros_shift.index)))
req_shift = req_shift.reindex(index=all_shifts, columns=all_dates, fill_value=0.0)
ros_shift = ros_shift.reindex(index=all_shifts, columns=all_dates, fill_value=0.0)
delt_shift = ros_shift - req_shift
for _df in (req_shift, ros_shift, delt_shift):
    _df.index.name = "Shift"
req_shift_total = add_total_row(req_shift)
ros_shift_total = add_total_row(ros_shift)
delt_shift_total = add_total_row(delt_shift)

# KPI row
# ==== KPI row (gradient cards) ====
def _fmt_int(n: float|int) -> str:
    try:
        v = int(round(float(n)))
        return f"{v:,}"
    except Exception:
        return str(n)

req_total  = _pretty_int(req_shift.values.sum())
ros_total  = _pretty_int(ros_shift.values.sum())
delta_total = _pretty_int(delt_shift.values.sum())

delta_class = "kpi-neutral"
if delta_total > 0:
    delta_class = "kpi-green"
elif delta_total < 0:
    delta_class = "kpi-red"

k1, k2, k3 = st.columns(3)

with k1:
    st.markdown(
        f"""
        <div class="kpi-card kpi-purple">
          <div class="kpi-label">Requested</div>
          <div class="kpi-value">{_fmt_int(req_total)}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

with k2:
    st.markdown(
        f"""
        <div class="kpi-card kpi-yellow">
          <div class="kpi-label">Rostered</div>
          <div class="kpi-value">{_fmt_int(ros_total)}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

with k3:
    st.markdown(
        f"""
        <div class="kpi-card {delta_class}">
          <div class="kpi-label">Delta</div>
          <div class="kpi-value">{_fmt_int(delta_total)}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

# =========================
# Renderers
# =========================
def render_dashboard(view_type: str,
                     req_shift: pd.DataFrame, ros_shift: pd.DataFrame, delt_shift: pd.DataFrame,
                     req_shift_total: pd.DataFrame, ros_shift_total: pd.DataFrame, delt_shift_total: pd.DataFrame):
    def _render_shiftwise():
        st.subheader("📈 Shift-wise View")
        sub1, sub2, sub3 = st.columns(3)
        with sub1:
            st.markdown("<div class='card'><h4>Requested – Shift-wise</h4>", unsafe_allow_html=True)
            st.dataframe(req_shift_total, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
        with sub2:
            st.markdown("<div class='card'><h4>Rostered – Shift-wise</h4>", unsafe_allow_html=True)
            st.dataframe(ros_shift_total, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
        with sub3:
            st.markdown("<div class='card'><h4>Delta – Shift-wise</h4>", unsafe_allow_html=True)
            st.dataframe(delt_shift_total, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

    def _render_intervalwise():
        st.subheader("⏱️ Interval-wise View")
        req_interval = transform_to_interval_view(req_shift)
        ros_interval = transform_to_interval_view(ros_shift)
        all_cols_i = sorted(list(set(req_interval.columns) | set(ros_interval.columns)))
        req_interval = req_interval.reindex(columns=all_cols_i, fill_value=0.0)
        ros_interval = ros_interval.reindex(columns=all_cols_i, fill_value=0.0)
        delt_interval = ros_interval - req_interval
        i1, i2, i3 = st.columns(3)
        with i1:
            st.markdown("<div class='card'><h4>Requested – Interval-wise</h4>", unsafe_allow_html=True)
            st.dataframe(add_total_row(req_interval), use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
        with i2:
            st.markdown("<div class='card'><h4>Rostered – Interval-wise</h4>", unsafe_allow_html=True)
            st.dataframe(add_total_row(ros_interval), use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
        with i3:
            st.markdown("<div class='card'><h4>Delta – Interval-wise</h4>", unsafe_allow_html=True)
            st.dataframe(add_total_row(delt_interval), use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

    if view_type == "Shift_Wise Delta View":
        _render_shiftwise()
    elif view_type == "Interval_Wise Delta View":
        _render_intervalwise()
    else:  # "Overall_Delta View"
        _render_shiftwise()
        st.markdown("---")
        _render_intervalwise()

# ==== Adapter for Plotter ====
class DBAdapter:
    def __init__(self, records_df, roster_df, is_admin: bool, allowed_centers: list[str], has_full_access: bool):
        self.records = records_df
        self.roster = roster_df
        self.is_admin = is_admin
        self.allowed_centers = set(allowed_centers or [])
        self.has_full_access = bool(has_full_access)

    def centers(self, include_overall=False):
        cs = _centers_union(self.records, self.roster)
        if not self.has_full_access:
            cs = [c for c in cs if c != "Overall" and c in self.allowed_centers]
        return [c for c in cs if c != "Overall"] if not include_overall else cs

    def languages(self, center):
        return _languages_union(self.records, self.roster, center)

    def pivot_requested(self, center, langs, start, end):
        return pivot_requested(self.records, center, langs, start, end)

    def _can_edit_center(self, center: str) -> bool:
        if not self.is_admin:
            return False
        if self.has_full_access:
            return True
        return center in self.allowed_centers

    def upsert_requested_cells(self, center, language, edits: dict):
        if not self._can_edit_center(center):
            st.error("You don't have edit permission for this center.")
            return 0
        if not edits:
            return 0
        rows = []
        for (shift, d), v in edits.items():
            rows.append({
                "Center": center,
                "Language": language,
                "Shift": shift,
                "Metric": "Requested",
                "Date": d,
                "Value": float(v),
            })
        df = pd.DataFrame(rows)
        if DUCKDB_FILE_NAME:
            local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
            if err or not local_db or not file_id:
                st.error(f"DuckDB issue: {err or 'File not found.'}")
                return 0
            return duckdb_upsert_records(local_db, file_id, DUCKDB_FILE_NAME, df)
        else:
            cur = load_parquet_from_drive(RECORDS_FILE)
            if cur.empty:
                save_parquet_to_drive(RECORDS_FILE, df)
                return len(df)
            keys = ["Center","Language","Shift","Date"]
            left = cur.merge(df[keys].drop_duplicates(), on=keys, how="left", indicator=True)
            keep = left["_merge"] == "left_only"
            base = cur[keep].copy()
            new_all = pd.concat([base, df], ignore_index=True)
            save_parquet_to_drive(RECORDS_FILE, new_all)
            return len(df)

# =========================
# PLOTTER (Desktop-like)
# =========================
import re as _re_pl
from datetime import date as _date_pl
if "plot_grid" not in st.session_state:     st.session_state.plot_grid = None
if "plot_center" not in st.session_state:   st.session_state.plot_center = None
if "plot_lang" not in st.session_state:     st.session_state.plot_lang = None
if "plot_range" not in st.session_state:    st.session_state.plot_range = (None, None)
if "plot_dirty" not in st.session_state:    st.session_state.plot_dirty = False

_TIME_LIKE = re.compile(r"^\d{4}-\d{4}(?:/\d{4}-\d{4})*$")

def _parse_ddmm_header_to_date(lbl: str, fallback_year: int) -> date | None:
    s = str(lbl).strip()
    for fmt in ("%d-%b-%Y", "%d-%b-%y", "%Y-%m-%d"):
        try:
            return pd.to_datetime(s, format=fmt).date()
        except Exception:
            pass
    try:
        return pd.to_datetime(f"{s}-{fallback_year}", format="%d-%b-%Y").date()
    except Exception:
        return None

def _ensure_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    has_total = df["Shift"].astype(str).str.lower().eq("total").any()
    if not has_total:
        new = df.copy()
        new.loc[len(new)] = ["Total"] + [0]*(df.shape[1]-1)
        return new
    return df

def _recalc_total(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return df
    body = df[~df["Shift"].astype(str).str.lower().eq("total")]
    sums = body.iloc[:, 1:].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=0)
    out = df.copy()
    mask_total = out["Shift"].astype(str).str.lower().eq("total")
    out.loc[mask_total, out.columns[1:]] = sums.values
    return out

def _grid_to_requested_df(grid_df: pd.DataFrame, from_year: int) -> pd.DataFrame:
    if grid_df is None or grid_df.empty: return pd.DataFrame()
    cols = list(grid_df.columns)
    if not cols or cols[0] != "Shift": return pd.DataFrame()
    parsed = []
    for h in cols[1:]:
        d = _parse_ddmm_header_to_date(h, from_year)
        parsed.append(d)
    keep = [i+1 for i, d in enumerate(parsed) if d is not None]
    out_dates = [d for d in parsed if d is not None]
    if not keep: return pd.DataFrame()
    body = grid_df[~grid_df["Shift"].astype(str).str.lower().eq("total")].copy()
    body_rows = []; idx_labels = []
    for _, r in body.iterrows():
        lbl = str(r["Shift"]).strip()
        if not lbl or not _TIME_LIKE.fullmatch(lbl):
            continue
        vals = []
        for j in keep:
            try:
                v = float(r.iloc[j])
            except Exception:
                v = 0.0
            vals.append(int(v) if float(v).is_integer() else float(v))
        body_rows.append(vals); idx_labels.append(lbl)
    if not body_rows: return pd.DataFrame()
    df = pd.DataFrame(body_rows, index=idx_labels, columns=out_dates)
    return df

def _plotter_load_grid(db, centre: str, lang: str, d_from: date, d_to: date) -> pd.DataFrame:
    req_df = db.pivot_requested(centre, [lang], d_from, d_to)
    req_df = dedup_requested_split_pairs(req_df)
    keep_cols = []
    for c in req_df.columns:
        cd = c.date() if hasattr(c, "date") else c
        if d_from <= cd <= d_to:
            keep_cols.append(c)
    keep_cols = sorted(keep_cols)
    req_df = req_df.reindex(columns=keep_cols, fill_value=0.0)
    headers = ["Shift"] + [pd.to_datetime(c).strftime("%d-%b") for c in keep_cols]
    grid = []
    for shift_label in req_df.index:
        row = [str(shift_label)]
        for c in keep_cols:
            v = req_df.at[shift_label, c]
            try:
                vf = float(v); row.append(int(vf) if vf.is_integer() else round(vf, 2))
            except Exception:
                row.append(0)
        grid.append(row)
    df = pd.DataFrame(grid, columns=headers)
    df = _ensure_total_row(df)
    df = _recalc_total(df)
    return df

def _plotter_interval_from_grid(grid_df: pd.DataFrame, from_year: int) -> pd.DataFrame:
    req = _grid_to_requested_df(grid_df, from_year)
    if req.empty:
        return pd.DataFrame(columns=["Interval"])
    req.index.name = "Shift"
    inter = transform_to_interval_view(req)
    inter = inter.reindex(index=range(24), fill_value=0.0)
    inter = add_total_row(inter)
    inter.index.name = "Interval"
    return inter

def render_plotter(db):
    st.subheader("🧮 Plotter (desktop workflow)")
    c1, c2, c3, c4, c5 = st.columns([1.2, 1.2, 1, 1, 1.6])
    with c1:
        centers = db.centers(include_overall=False) or []
        centre = st.selectbox("Center", centers, index=0 if centers else None, key="plot_center_sel")
    with c2:
        langs = db.languages(centre) if centre else []
        lang = st.selectbox("Language", langs, index=0 if langs else None, key="plot_lang_sel")
    with c3:
        d_from = st.date_input("From", value=st.session_state.get("plot_from") or date.today(), key="plot_from")
    with c4:
        d_to = st.date_input("To", value=st.session_state.get("plot_to") or date.today(), key="plot_to")

    def _reload_grid():
        if not centre or not lang: return
        with overlay("Preparing editable grid…"):
            grid = _plotter_load_grid(db, centre, lang, d_from, d_to)
            st.session_state.plot_grid = grid
            st.session_state.plot_center = centre
            st.session_state.plot_lang = lang
            st.session_state.plot_range = (d_from, d_to)
            st.session_state.plot_dirty = False

    if (st.session_state.plot_grid is None or
        st.session_state.plot_center != centre or
        st.session_state.plot_lang != lang or
        st.session_state.plot_range != (d_from, d_to)):
        _reload_grid()

    with c5:
        b1, b2, b3, b4 = st.columns([1,1,1,1.2])
        if b1.button("Add Date"):
            if st.session_state.plot_grid is not None and not st.session_state.plot_grid.empty:
                headers = list(st.session_state.plot_grid.columns)
                base_year = d_from.year
                existing = set()
                for h in headers[1:]:
                    dd = _parse_ddmm_header_to_date(h, base_year)
                    if dd: existing.add(dd)
                wanted = pd.date_range(d_from, d_to, freq="D").date
                to_add = [d for d in wanted if d not in existing]
                if to_add:
                    new_headers = headers + [d.strftime("%d-%b") for d in to_add]
                    g = st.session_state.plot_grid.copy()
                    for nh in new_headers:
                        if nh not in g.columns: g[nh] = 0
                    head = ["Shift"]
                    pairs = []
                    for h in g.columns[1:]:
                        dd = _parse_ddmm_header_to_date(h, base_year)
                        if dd: pairs.append((h, dd))
                    pairs.sort(key=lambda x: x[1]); head += [h for h,_ in pairs]
                    g = g[head]
                    g = _ensure_total_row(g)
                    g = _recalc_total(g)
                    st.session_state.plot_grid = g
                    st.session_state.plot_dirty = True
                else:
                    st.info("No new dates in range.")
            else:
                _reload_grid()

        b2.button("Manage Shifts", help="(Open shifts screen in your web app)")
        b3.button("Manage Languages", help="(Open languages screen in your web app)")

        can_edit_here = db._can_edit_center(centre)
        if b4.button("Save", type="primary", disabled=not can_edit_here):
            with overlay("Saving Requested edits…"):
                grid = st.session_state.plot_grid
                if grid is None or grid.empty or grid.shape[1] < 2:
                    st.warning("Nothing to save.")
                else:
                    req_df = _grid_to_requested_df(grid, d_from.year)
                    edits = {}
                    for sh, row in req_df.iterrows():
                        for d, v in row.items():
                            if float(v) != 0.0:
                                edits[(sh, d)] = float(v)
                    if edits:
                        db.upsert_requested_cells(centre, lang, edits)
                        st.success(f"Saved {len(edits)} cells.")
                        st.session_state.plot_dirty = False
                    else:
                        st.info("No non-zero changes to save.")

    st.caption("Editing: " + ("ENABLED" if db._can_edit_center(centre) else "VIEW ONLY"))
    st.markdown("---")

    st.caption("Shift-wise Requested (editable)")
    if st.session_state.plot_grid is None:
        st.info("Select Center, Language and date range to load.")
        return
    cfg = {"Shift": st.column_config.TextColumn("Shift", disabled=True)}
    edited = st.data_editor(
        st.session_state.plot_grid,
        num_rows="dynamic",
        use_container_width=True,
        column_config=cfg,
        key="plot_editor",
    )
    if not edited.empty:
        if "Shift" in edited.columns:
            edited.loc[edited["Shift"].astype(str).str.lower() == "total", "Shift"] = "Total"
        mask_total = edited["Shift"].astype(str).str.lower() == "total"
        if mask_total.any():
            cols = [c for c in edited.columns if c != "Shift"]
            edited.loc[mask_total, cols] = 0
        edited = _ensure_total_row(edited)
        edited = _recalc_total(edited)

    st.session_state.plot_grid = edited
    st.session_state.plot_dirty = True
    inter = _plotter_interval_from_grid(edited, d_from.year)
    if inter.empty:
        st.info("No interval view for current grid.")
    else:
        st.caption("Interval-wise Requested (auto)")
        st.dataframe(inter, use_container_width=True)

    st.markdown(
        f"<div style='text-align:right;color:#cbd5e1'>"
        f"{'Unsaved changes' if st.session_state.plot_dirty else 'Up to date'}"
        f"</div>",
        unsafe_allow_html=True
    )

def render_roster(roster: pd.DataFrame, center: str, start: date, end: date):
    # --- Admin center toggle (read-only guard)
    center_edit_ok = center_can_edit(center) if center else False
    if not center_edit_ok:
        st.info("✋ This center is currently **read-only** (edit access disabled by Admin).")
    # Normalize/validate schema first
    roster = _ensure_roster_schema(roster)

    # Fast guard: show message if required columns are missing
    if roster.empty or "Date" not in roster.columns:
        st.warning("Roster table has no 'Date' column. Please import a valid roster with a 'Roster' sheet containing at least 'Date' and 'Shift'.")
        return

    def _roster_set_busy():
        st.session_state.busy_roster = True
    disabled = st.session_state.busy_roster

    r1, r2, r3, r4, r5 = st.columns([1,1,1,1.3,1.2])
    with r1:
        lobs = roster_lobs(roster, center, start, end)
        lob_sel = st.selectbox("LOB", ["(All)"] + lobs, disabled=disabled)
        lob = None if lob_sel in ("", "(All)") else lob_sel
    with r2:
        langs_lob = roster_languages_for_lob(roster, center, lob, start, end)
        rlang_sel = st.selectbox("Language", ["(All)"] + langs_lob, key="rlang_sel",
                                 on_change=_roster_set_busy, disabled=disabled)
        rlang = None if rlang_sel in ("", "(All)") else rlang_sel
    with r3:
        codes = roster_nonshift_codes(roster, center, start, end, lob, rlang)
        nonshift_sel = st.selectbox("Non-Shift", ["(All)"] + codes, disabled=disabled)
    with r4:
        shifts_list = roster_distinct_shifts(roster, center, start, end, lob, rlang)
        shift_disabled = disabled or (nonshift_sel not in ("", "(All)"))
        shift_sel = st.selectbox("Shift", ["(All)"] + shifts_list, disabled=shift_disabled)
    with r5:
        ids_text = st.text_input("Genesys IDs (comma/space)", disabled=disabled)
        ids = [x.strip() for x in re.findall(r"[A-Za-z0-9_]+", ids_text)] if ids_text else None

    # Use full-screen overlay for the heavy filtering/rendering block
    with overlay("Loading roster…"):
        dfr = roster.copy()

        # Date filter (safe: 'Date' exists by now)
        dfr = dfr[dfr["Date"].between(start, end)]

        # Optional filters
        if center and center != "Overall":
            if "Center" in dfr.columns:
                dfr = dfr[dfr["Center"] == center]
        if lob and "LOB" in dfr.columns:
            dfr = dfr[dfr["LOB"] == lob]
        if rlang and "Language" in dfr.columns:
            dfr = dfr[dfr["Language"] == rlang]
        if ids and "AgentID" in dfr.columns:
            dfr = dfr[dfr["AgentID"].isin(ids)]

        # Normalize Shift again post-filter (defensive)
        if not dfr.empty and "Shift" in dfr.columns:
            dfr["Shift"] = dfr["Shift"].apply(lambda s: (_normalize_shift_label(s) or str(s).strip().upper()))

        # Apply Non-Shift vs Shift filters
        if not dfr.empty and "Shift" in dfr.columns:
            time_like = re.compile(r"^\d{4}-\d{4}(?:/\d{4}-\d{4})*$")
            single_pat = re.compile(r"^\d{4}-\d{4}$")
            if nonshift_sel not in ("", "(All)"):
                dfr = dfr[(~dfr["Shift"].astype(str).str.match(time_like)) &
                          (dfr["Shift"].astype(str).str.upper() == nonshift_sel.upper())]
            elif shift_sel not in ("", "(All)"):
                chosen_is_single = bool(single_pat.fullmatch(shift_sel))
                def keep_shift(s: str) -> bool:
                    s = str(s).strip()
                    if not time_like.fullmatch(s): return False
                    if s == shift_sel: return True
                    if chosen_is_single and "/" in s:
                        parts = [p.strip() for p in s.split("/") if p.strip()]
                        return shift_sel in parts
                    return False
                dfr = dfr[dfr["Shift"].apply(keep_shift)]

        # Reorder and format for display
        static_cols = [c for c in [
            "AgentID","AgentName","TLName","Status","WorkMode","Center","Location",
            "Language","SecondaryLanguage","LOB","FTPT","Shift"
        ] if c in dfr.columns]

        if not dfr.empty:
            long_df = dfr[["Date"] + static_cols].copy()
            long_df["Date"] = pd.to_datetime(long_df["Date"], errors="coerce").dt.strftime("%d-%b-%y")
        else:
            long_df = pd.DataFrame(columns=["Date"] + static_cols)

        st.caption("Roster (long format)")
        st.dataframe(long_df, use_container_width=True)

    st.session_state.busy_roster = False

    # ======= Edit / Update roster (dialog) =======
    @st.dialog("Edit / Update Roster")
    def roster_edit_dialog(roster_df: pd.DataFrame):
        if not can_edit_this_center:
            st.error("You don't have edit permission for this center.")
            return
        mode = st.radio("Edit mode", ["Single edit", "Bulk edit (upload)"], horizontal=True)
        if mode == "Single edit":
            agent = st.text_input("Genesys ID (single)")
            c1, c2 = st.columns(2)
            with c1:
                d_from = st.date_input("From date", value=start)
            with c2:
                d_to = st.date_input("To date", value=min(end, start + timedelta(days=6)))
            # enforce 7-day window
            if (d_to - d_from).days > 6:
                st.warning("Maximum 7 days allowed. Adjusting.")
                d_to = d_from + timedelta(days=6)
            new_shift = st.text_input("New Shift (e.g., 0900-1800 or 0900-1300/1400-1800)")
            if st.button("Apply", type="primary", use_container_width=True, disabled=not (agent and new_shift)):
                norm = _normalize_shift_label(new_shift)
                if not norm:
                    st.error("Invalid shift format.")
                else:
                    with overlay("Validating & submitting…"):
                        try:
                            # validations
                            werr = validate_window(d_from, d_to)
                            if werr:
                                st.error(werr);
                                return

                            # build items df for [AgentID × date range], single Shift applied per day
                            dates = pd.date_range(d_from, d_to, freq="D").date
                            # validate shift value with admin-managed codes
                            if DUCKDB_FILE_NAME:
                                local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
                                if err:
                                    st.error(err);
                                    return
                                con = duckdb.connect(local_db, read_only=True)
                            else:
                                con = None

                            ok, reason = validate_shift_value(norm, con)
                            if con is not None:
                                try:
                                    con.close()
                                except:
                                    pass
                            if not ok:
                                st.error(f"Shift invalid: {reason}")
                                return

                            items = []
                            for d in dates:
                                items.append({
                                    "AgentID": str(agent).strip(),
                                    "Date": d,
                                    "Shift": norm,
                                    "TLName": None,
                                    "Status": None,
                                    "Language": None,
                                    "LOB": None,
                                    "WorkMode": None
                                })
                            df_items = pd.DataFrame(items)

                            req_id = submit_roster_request(
                                mode="Single",
                                center=center,
                                uploaded_by=user.get("email", ""),
                                from_date=d_from,
                                to_date=d_to,
                                items_df=df_items
                            )
                            st.success(f"Submitted for approval. Request ID: {req_id}")
                        except Exception as e:
                            st.error(f"Submission failed: {e}")
        else:
            st.caption("Upload a roster file to bulk replace (validation rules TBD).")
            up = st.file_uploader("Roster (.xlsx) with 'Roster' sheet", type=["xlsx"], key="bulk_roster_up")
            if st.button("Upload & Validate (Pending)", type="primary", disabled=not up):
                with overlay(f"Validating bulk file for {center}…"):
                    raw = read_roster_sheet(up.read())
                    if raw.empty:
                        st.error("No 'Roster' sheet found or it's empty.")
                    else:
                        # Ensure schema & coerce
                        df = _ensure_roster_schema(raw)
                        if "Date" in df.columns:
                            df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date

                        # 1) window
                        werr = validate_window(start, end)
                        if werr:
                            st.error(werr);
                            st.stop()

                        # 2) exact date columns (for per-day columns like '06-Oct', '07-Oct', ...)
                        #    If your bulk template uses one 'Date' column instead, we accept that too.
                        wanted_dates = pd.date_range(start, end, freq="D").date
                        per_day_headers = []
                        for c in df.columns:
                            # detect date-like headers (e.g., 06-Oct, 2024-10-06, 06/10, etc.)
                            dtry = _parse_ddmm_header_to_date(c, start.year)
                            if dtry: per_day_headers.append((c, dtry))

                        if per_day_headers:
                            header_dates = sorted([d for _, d in per_day_headers])
                            if list(header_dates) != list(wanted_dates):
                                st.error(
                                    "Date-wise columns must exactly match the selected From–To range (no extra/missing days).")
                                st.stop()

                        # 3) row validations
                        errors = []
                        # get admin shift codes for validation
                        if DUCKDB_FILE_NAME:
                            local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
                            con = duckdb.connect(local_db, read_only=True) if not err else None
                        else:
                            con = None

                        def _val_shift(x):
                            ok, why = validate_shift_value(_normalize_shift_label(x) or x, con)
                            return (ok, why)

                        for idx, r in df.iterrows():
                            # Scope check
                            if str(r.get("Center", "")).strip() not in (center, ""):
                                errors.append((idx, "Center must equal your center"));
                                continue

                            # Allowed-values
                            errors += [(idx, msg) for msg in validate_allowed_values(r.to_dict())]

                            # Shift validation:
                            if per_day_headers:
                                for col, d in per_day_headers:
                                    val = r.get(col, "")
                                    ok, why = _val_shift(val)
                                    if not ok:
                                        errors.append((idx, f"{col}: {why}"))
                            else:
                                val = r.get("Shift", "")
                                ok, why = _val_shift(val)
                                if not ok:
                                    errors.append((idx, f"Shift: {why}"))

                        if con is not None:
                            try:
                                con.close()
                            except:
                                pass

                        if errors:
                            st.error(f"Validation errors: {len(errors)}. Fix and retry.")
                            # optional: show first 30 errors
                            show = errors[:30]
                            st.code("\n".join([f"Row {i}: {msg}" for i, msg in show]))
                            st.stop()

                        # Build items df for submission
                        items = []
                        if per_day_headers:
                            for _, r in df.iterrows():
                                agent = str(r.get("AgentID", "")).strip()
                                for col, d in per_day_headers:
                                    sh = _normalize_shift_label(r.get(col, "")) or str(r.get(col, "")).strip().upper()
                                    items.append({
                                        "AgentID": agent, "Date": d, "Shift": sh,
                                        "TLName": r.get("TLName"), "Status": r.get("Status"),
                                        "Language": r.get("Language"), "LOB": r.get("LOB"),
                                        "WorkMode": r.get("WorkMode")
                                    })
                        else:
                            # single 'Date' + 'Shift' format
                            for _, r in df.iterrows():
                                agent = str(r.get("AgentID", "")).strip()
                                d = r.get("Date")
                                if d is None or pd.isna(d): continue
                                sh = _normalize_shift_label(r.get("Shift", "")) or str(
                                    r.get("Shift", "")).strip().upper()
                                items.append({
                                    "AgentID": agent, "Date": d, "Shift": sh,
                                    "TLName": r.get("TLName"), "Status": r.get("Status"),
                                    "Language": r.get("Language"), "LOB": r.get("LOB"),
                                    "WorkMode": r.get("WorkMode")
                                })

                        df_items = pd.DataFrame(items)
                        if df_items.empty:
                            st.warning("Nothing to submit after validation.");
                            st.stop()

                        req_id = submit_roster_request(
                            mode="Bulk",
                            center=center,
                            uploaded_by=user.get("email", ""),
                            from_date=start,
                            to_date=end,
                            items_df=df_items
                        )
                        st.success(f"Submitted for approval. Request ID: {req_id}")

    st.button("Edit / Update Roster", on_click=lambda: roster_edit_dialog(roster),
              type="primary", disabled=not (can_edit_this_center and center_edit_ok))

def render_admin_panel(all_centers_no_overall: list[str], ACL: dict, user_email: str):
    st.subheader("🛡️ Admin")

    # ===== Submissions queue =====
    st.markdown("### Pending Submissions")
    try:
        local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
        if err or not local_db or not file_id:
            st.info("DuckDB not configured; submissions view unavailable.")
        else:
            con = duckdb.connect(local_db, read_only=True)
            _ensure_duckdb_schema(con)
            pending = con.execute("""
                SELECT request_id, mode, center, uploaded_by, from_date, to_date, status, reason, created_at
                FROM submissions
                WHERE status='Pending'
                ORDER BY created_at DESC
                LIMIT 200
            """).df()
            con.close()

            if pending.empty:
                st.caption("No pending submissions.")
            else:
                st.dataframe(pending, use_container_width=True, height=280)

                with st.form("admin_action"):
                    rid = st.text_input("Request ID")
                    action = st.selectbox("Action", ["Approve","Deny"])
                    reason = st.text_input("Reason (for Deny)")
                    submitted = st.form_submit_button("Apply", type="primary")
                if submitted and rid.strip():
                    with overlay(f"{action} request…"):
                        try:
                            if action == "Approve":
                                admin_approve_request(rid.strip(), user_email)
                                st.success(f"Approved {rid}")
                            else:
                                admin_deny_request(rid.strip(), user_email, reason)
                                st.warning(f"Denied {rid}")
                            st.experimental_rerun()
                        except Exception as e:
                            st.error(f"Action failed: {e}")
    except Exception as e:
        st.error(f"Submissions section error: {e}")

    st.markdown("---")

    # ===== Center Edit Access =====
    st.markdown("### Center Edit Access (ON/OFF)")
    cA, cB = st.columns([1.1,1.9])
    with cA:
        center_sel = st.selectbox("Center", all_centers_no_overall, index=0 if all_centers_no_overall else None, key="perm_center_sel")
    with cB:
        if center_sel:
            try:
                local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
                if err or not local_db or not file_id:
                    st.info("DuckDB not configured; skipping permissions.")
                else:
                    con = duckdb.connect(local_db, read_only=False)
                    _ensure_duckdb_schema(con)
                    row = con.execute("SELECT can_edit FROM center_permissions WHERE center=?", [center_sel]).df()
                    can_edit = True if row.empty else bool(row.iloc[0]["can_edit"])
                    toggle = st.toggle("Can Edit", value=can_edit)
                    if st.button("Save Access", type="primary"):
                        con.execute("""
                            INSERT INTO center_permissions (center, can_edit) VALUES (?, ?)
                            ON CONFLICT (center) DO UPDATE SET can_edit=excluded.can_edit
                        """, [center_sel, toggle])
                        _audit_log(con, "CenterAccessToggle", center_sel, user_email, f"can_edit={toggle}")
                        st.success("Saved.")
                    con.close()
            except Exception as e:
                st.error(f"Center access error: {e}")

    st.markdown("---")

    # ===== Shift Codes Manager =====
    st.markdown("### Shift Codes (Admin-managed)")
    st.caption("Use this to add non-time codes (e.g., FS) or time-bound codes with constraints.")
    try:
        local_db, file_id, err = _download_duckdb_rw(DUCKDB_FILE_NAME)
        if err or not local_db or not file_id:
            st.info("DuckDB not configured; skipping shift codes.")
        else:
            con = duckdb.connect(local_db, read_only=False)
            _ensure_duckdb_schema(con)
            df = con.execute("SELECT code, label, category, is_time_code, min_duration_minutes, notes, is_enabled FROM shift_codes ORDER BY code").df()
            # Editable grid
            edited = st.data_editor(
                df if not df.empty else pd.DataFrame(columns=["code","label","category","is_time_code","min_duration_minutes","notes","is_enabled"]),
                num_rows="dynamic",
                use_container_width=True,
                key="shift_codes_editor"
            )
            if st.button("Save Shift Codes", type="primary"):
                with overlay("Saving shift codes…"):
                    con.execute("DELETE FROM shift_codes")
                    if not edited.empty:
                        con.register("codes_src", edited)
                        con.execute("""
                            INSERT INTO shift_codes
                            SELECT UPPER(code), label, category, COALESCE(is_time_code, FALSE),
                                   NULLIF(min_duration_minutes, NULL), notes, COALESCE(is_enabled, TRUE)
                            FROM codes_src
                            WHERE TRIM(code) <> ''
                        """)
                    _audit_log(con, "ShiftCodesSave", "(all)", user_email, f"rows={0 if edited is None else len(edited)}")
                    st.success("Saved shift codes.")
            con.close()
    except Exception as e:
        st.error(f"Shift Codes error: {e}")

    st.markdown("---")

    # ===== Existing Editors management (your original UI) =====
    st.markdown("### Center Editors (Legacy email-based)")
    colA, colB = st.columns([1.2, 1.8])
    with colA:
        center_sel2 = st.selectbox("Center (Editors)", all_centers_no_overall, index=0 if all_centers_no_overall else None)
    if not center_sel2:
        st.info("No centers available.")
        return
    editors = sorted(set((ACL.get(center_sel2, {}).get("editors") or [])), key=str.lower)
    with colB:
        st.write("**Current Editors**")
        if editors:
            for i, em in enumerate(editors):
                c1, c2 = st.columns([6,1])
                with c1: st.write(em)
                with c2:
                    if st.button("❌", key=f"rm_{center_sel2}_{i}"):
                        new = set(editors); new.discard(em)
                        ACL.setdefault(center_sel2, {})["editors"] = sorted(new, key=str.lower)
                        _save_acl_to_drive(ACL, DRIVE_FOLDER_ID or None)
                        st.success(f"Removed {em}")
                        st.experimental_rerun()
        else:
            st.caption("_No explicit editors — only ‘meesho’ or email-matched users can edit._")
        st.markdown("---")
        new_email = st.text_input("Grant edit access (email)")
        if st.button("➕ Add Editor"):
            if not new_email or "@" not in new_email:
                st.warning("Enter a valid email.")
            else:
                ACL.setdefault(center_sel2, {}).setdefault("editors", [])
                if new_email.lower() not in map(str.lower, ACL[center_sel2]["editors"]):
                    ACL[center_sel2]["editors"].append(new_email)
                    _save_acl_to_drive(ACL, DRIVE_FOLDER_ID or None)
                    st.success(f"Added {new_email}")
                    st.experimental_rerun()
                else:
                    st.info("Already present.")

# ------- Persistent tab switcher -------
db = DBAdapter(records, roster, is_admin=is_admin, allowed_centers=allowed_centers, has_full_access=has_full_access)
tabs_base = ["Dashboard", "Plotter", "Roster"]
if is_admin: tabs_base.append("Admin")
tab_choice = st.radio("View", tabs_base, horizontal=True, key="active_tab")

if tab_choice == "Dashboard":
    render_dashboard(
        view_type,
        req_shift, ros_shift, delt_shift,
        req_shift_total, ros_shift_total, delt_shift_total
    )
elif tab_choice == "Plotter":
    render_plotter(db)
elif tab_choice == "Roster":
    render_roster(roster, center, start, end)
else:
    all_centers_no_overall = [c for c in _centers_union(records, roster) if c and c != "Overall"]
    render_admin_panel(all_centers_no_overall, ACL, user.get("email",""))
